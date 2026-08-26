"""
Fill empty Size cells on Size References from ProductExport.

Matches the last numeric (...) group in SKU Value to ProductExport UID.
Re-run anytime after adding Size References rows; existing Size values are left alone.
Close Configuration Workbook.xlsx in Excel before running.
"""

from __future__ import annotations

import re
import sys
from pathlib import Path

import openpyxl

_WAREHOUSE = Path(__file__).resolve().parents[2]
if str(_WAREHOUSE) not in sys.path:
    sys.path.insert(0, str(_WAREHOUSE))
from shared import paths as wh  # noqa: E402

WORKBOOK_PATH = wh.queue_config_workbook_path()
PRODUCT_EXPORT_PATH = wh.data_archive_dir() / "Queue_ProductExport.xlsx"
if not PRODUCT_EXPORT_PATH.exists():
    # Prefer shared CSV PE via openpyxl only for xlsx; archive holds Queue xlsx
    PRODUCT_EXPORT_PATH = wh.data_archive_dir() / "PO_ProductExport.xlsx"

SIZE_REFERENCES_SHEET = "Size References"
PRODUCT_EXPORT_SHEET = "staff"

SKU_COL = 2  # B
SIZE_COL = 7  # G
UID_COL = 1  # A
EXPORT_SIZE_COL = 15  # O

UID_IN_BRACKETS = re.compile(r"\((\d+)\)")


def is_blank(value) -> bool:
    return value is None or str(value).strip() == ""


def extract_uid(sku) -> str | None:
    if sku is None:
        return None
    matches = UID_IN_BRACKETS.findall(str(sku))
    return matches[-1] if matches else None


def load_uid_to_size(path: Path) -> dict[str, str]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[PRODUCT_EXPORT_SHEET]
        uid_to_size: dict[str, str] = {}
        # Rows 1-2 are headers / placeholders
        for row in ws.iter_rows(min_row=3, values_only=True):
            uid = row[UID_COL - 1]
            size = row[EXPORT_SIZE_COL - 1]
            if uid is None or is_blank(size):
                continue
            uid_to_size[str(uid).strip()] = str(size).strip()
        return uid_to_size
    finally:
        wb.close()


def fill_sizes(workbook_path: Path, uid_to_size: dict[str, str]) -> dict[str, int]:
    wb = openpyxl.load_workbook(workbook_path)
    try:
        ws = wb[SIZE_REFERENCES_SHEET]
        stats = {
            "filled": 0,
            "already_had_size": 0,
            "no_numeric_uid": 0,
            "uid_not_found": 0,
            "blank_sku": 0,
        }

        for row_idx in range(2, ws.max_row + 1):
            sku = ws.cell(row=row_idx, column=SKU_COL).value
            size_cell = ws.cell(row=row_idx, column=SIZE_COL)

            if not is_blank(size_cell.value):
                stats["already_had_size"] += 1
                continue

            if is_blank(sku):
                stats["blank_sku"] += 1
                continue

            uid = extract_uid(sku)
            if uid is None:
                stats["no_numeric_uid"] += 1
                continue

            size = uid_to_size.get(uid)
            if size is None:
                stats["uid_not_found"] += 1
                continue

            size_cell.value = size
            stats["filled"] += 1

        wb.save(workbook_path)
        return stats
    finally:
        wb.close()


def main() -> int:
    if not WORKBOOK_PATH.exists():
        print(f"Missing: {WORKBOOK_PATH}", file=sys.stderr)
        return 1
    if not PRODUCT_EXPORT_PATH.exists():
        print(f"Missing: {PRODUCT_EXPORT_PATH}", file=sys.stderr)
        return 1

    print(f"Loading sizes from {PRODUCT_EXPORT_PATH.name}...")
    uid_to_size = load_uid_to_size(PRODUCT_EXPORT_PATH)
    print(f"  {len(uid_to_size)} UID -> Size entries")

    print(f"Updating {WORKBOOK_PATH.name} / {SIZE_REFERENCES_SHEET}...")
    try:
        stats = fill_sizes(WORKBOOK_PATH, uid_to_size)
    except PermissionError:
        print(
            f"Cannot save {WORKBOOK_PATH.name}. Close it in Excel and try again.",
            file=sys.stderr,
        )
        return 1

    print("Done.")
    print(f"  Filled:            {stats['filled']}")
    print(f"  Already had size:  {stats['already_had_size']}")
    print(f"  No numeric UID:    {stats['no_numeric_uid']}")
    print(f"  UID not found:     {stats['uid_not_found']}")
    print(f"  Blank SKU:         {stats['blank_sku']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
