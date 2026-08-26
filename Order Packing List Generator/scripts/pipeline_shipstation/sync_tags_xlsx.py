"""Sync Data/ShipStation Tags.xlsx with live ShipStation tags.

Adds missing tags (name + tagId), updates Tag IDs when they change,
and leaves Process No columns untouched. Does not delete obsolete rows.
"""

from __future__ import annotations

import argparse
import shutil
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .client import ShipStationClient, ShipStationError
from .credentials import load_shipstation_credentials

PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
_WAREHOUSE = PROJECT_ROOT.parent
if str(_WAREHOUSE) not in sys.path:
    sys.path.insert(0, str(_WAREHOUSE))
from shared import paths as wh  # noqa: E402

DEFAULT_XLSX_PATH = wh.shipstation_tags_path()
SHEET_NAME = "Tags"

COL_SR = 1
COL_NAME = 2
COL_ID = 3
# Process No columns D–H are left as-is for existing rows and blank for new ones.

HEADERS = [
    "Sr. No.",
    "Tag Name",
    "Tag ID",
    "Process No - 1st Shift",
    "Process No - 2nd Shift",
    "Process No - 3rd Shift",
    "Process No - 4th Shift",
    "Process No - 5th Shift",
]


@dataclass
class SyncResult:
    live_count: int = 0
    excel_count_before: int = 0
    excel_count_after: int = 0
    updated_ids: list[tuple[str, int | None, int]] = field(default_factory=list)
    added: list[tuple[str, int]] = field(default_factory=list)
    obsolete: list[tuple[str, int | None]] = field(default_factory=list)
    dry_run: bool = False
    path: Path | None = None
    backup_path: Path | None = None


def _name_key(name: str) -> str:
    return str(name or "").strip().casefold()


def _as_int(value: Any) -> int | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        return int(float(text)) if "." in text else int(text)
    except (TypeError, ValueError):
        return None


def _ensure_headers(ws: Worksheet) -> None:
    for col, header in enumerate(HEADERS, start=1):
        current = ws.cell(1, col).value
        if current is None or str(current).strip() == "":
            ws.cell(1, col).value = header


def _read_excel_rows(ws: Worksheet) -> dict[str, tuple[int, str, int | None]]:
    """Map casefolded tag name -> (row, display name, tag id)."""
    by_name: dict[str, tuple[int, str, int | None]] = {}
    for row in range(2, ws.max_row + 1):
        name_raw = ws.cell(row, COL_NAME).value
        if name_raw is None or str(name_raw).strip() == "":
            continue
        name = str(name_raw).strip()
        key = _name_key(name)
        by_name[key] = (row, name, _as_int(ws.cell(row, COL_ID).value))
    return by_name


def _next_sr_no(ws: Worksheet) -> int:
    last = 0
    for row in range(2, ws.max_row + 1):
        n = _as_int(ws.cell(row, COL_SR).value)
        if n is not None and n > last:
            last = n
    return last + 1


def _create_empty_workbook(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    for col, header in enumerate(HEADERS, start=1):
        ws.cell(1, col).value = header
    wb.save(path)


def sync_shipstation_tags_xlsx(
    xlsx_path: str | Path | None = None,
    *,
    client: ShipStationClient | None = None,
    dry_run: bool = False,
    backup: bool = True,
) -> SyncResult:
    """
    Fetch ShipStation tags and sync into the Tags sheet.

    - Adds tags present in ShipStation but missing from Excel.
    - Updates Tag ID when an existing name's id changed.
    - Does not remove Excel rows that are no longer in ShipStation
      (reported as obsolete).
    - Leaves Process No columns unchanged for existing rows; blank for new rows.
    """
    path = Path(xlsx_path) if xlsx_path else DEFAULT_XLSX_PATH
    result = SyncResult(dry_run=dry_run, path=path)

    if not path.is_file():
        if dry_run:
            raise FileNotFoundError(f"Tags workbook not found: {path}")
        _create_empty_workbook(path)

    ss = client or ShipStationClient(load_shipstation_credentials())
    live_tags = ss.list_tags()
    live_by_name = {
        _name_key(t["name"]): {"tagId": int(t["tagId"]), "name": str(t["name"]).strip()}
        for t in live_tags
        if t.get("name") is not None and str(t.get("name")).strip()
    }
    result.live_count = len(live_by_name)

    wb = load_workbook(path)
    if SHEET_NAME in wb.sheetnames:
        ws = wb[SHEET_NAME]
    else:
        ws = wb.active
        ws.title = SHEET_NAME
    _ensure_headers(ws)

    excel_by_name = _read_excel_rows(ws)
    result.excel_count_before = len(excel_by_name)

    for key, live in live_by_name.items():
        if key not in excel_by_name:
            continue
        row, name, excel_id = excel_by_name[key]
        live_id = live["tagId"]
        if excel_id != live_id:
            result.updated_ids.append((name, excel_id, live_id))
            if not dry_run:
                ws.cell(row, COL_ID).value = live_id

    next_sr = _next_sr_no(ws)
    for key, live in sorted(live_by_name.items(), key=lambda kv: kv[1]["tagId"]):
        if key in excel_by_name:
            continue
        name = live["name"]
        tag_id = live["tagId"]
        result.added.append((name, tag_id))
        if not dry_run:
            row = ws.max_row + 1
            ws.cell(row, COL_SR).value = next_sr
            ws.cell(row, COL_NAME).value = name
            ws.cell(row, COL_ID).value = tag_id
            next_sr += 1

    for key, (row, name, excel_id) in excel_by_name.items():
        if key not in live_by_name:
            result.obsolete.append((name, excel_id))

    result.excel_count_after = result.excel_count_before + len(result.added)

    if dry_run:
        return result

    if backup and path.is_file() and (result.added or result.updated_ids):
        bak = path.with_suffix(path.suffix + ".bak")
        shutil.copy2(path, bak)
        result.backup_path = bak

    try:
        wb.save(path)
    except PermissionError as exc:
        raise PermissionError(
            f"Cannot write {path.name} — close it in Excel and try again."
        ) from exc
    return result


def _print_report(result: SyncResult) -> None:
    path = result.path or DEFAULT_XLSX_PATH
    mode = "DRY RUN - no changes written" if result.dry_run else "Synced"
    print(f"{mode}: {path}")
    print(f"  ShipStation tags : {result.live_count}")
    print(f"  Excel before     : {result.excel_count_before}")
    print(f"  Excel after      : {result.excel_count_after}")
    print(f"  Tag IDs updated  : {len(result.updated_ids)}")
    print(f"  Tags added       : {len(result.added)}")
    print(f"  Obsolete in Excel: {len(result.obsolete)} (kept; not deleted)")
    if result.backup_path:
        print(f"  Backup           : {result.backup_path}")
    if result.updated_ids:
        print("\nUpdated Tag IDs:")
        for name, old_id, new_id in result.updated_ids:
            print(f"  - {name}: {old_id} -> {new_id}")
    if result.added:
        print("\nAdded tags:")
        for name, tag_id in result.added:
            print(f"  - {name} ({tag_id})")
    if result.obsolete:
        print("\nObsolete in Excel (still in file):")
        for name, tag_id in result.obsolete:
            print(f"  - {name} ({tag_id})")


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description=(
            "Sync Data/ShipStation Tags.xlsx with current ShipStation tags "
            "(add missing names/IDs, update changed IDs)."
        )
    )
    parser.add_argument(
        "--xlsx",
        type=Path,
        default=DEFAULT_XLSX_PATH,
        help=f"Path to ShipStation Tags workbook (default: {DEFAULT_XLSX_PATH})",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Report what would change without writing the file.",
    )
    parser.add_argument(
        "--no-backup",
        action="store_true",
        help="Do not write a .xlsx.bak before saving.",
    )
    args = parser.parse_args(argv)

    try:
        result = sync_shipstation_tags_xlsx(
            args.xlsx,
            dry_run=args.dry_run,
            backup=not args.no_backup,
        )
    except (ShipStationError, FileNotFoundError, ValueError, PermissionError) as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 1

    _print_report(result)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
