"""Look up process numbers from Data/ShipStation Tags.xlsx by tag + shift."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .sync_tags_xlsx import DEFAULT_XLSX_PATH, SHEET_NAME, _as_int, _name_key

# GUI shift labels -> 1-based column index in Tags sheet.
SHIFT_PROCESS_COLS: dict[str, int] = {
    "1st": 4,  # Process No - 1st Shift
    "2nd": 5,
    "3rd": 6,
    "4th": 7,
    "5th": 8,
}


def _normalize_shift(shift_label: str) -> str:
    text = (shift_label or "").strip()
    # Accept "1st", "1st Shift", etc.
    for key in SHIFT_PROCESS_COLS:
        if text.casefold().startswith(key.casefold()):
            return key
    return text


def _cell_to_process_str(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    as_int = _as_int(value)
    if as_int is not None:
        return str(as_int)
    return text


def lookup_process_number(
    *,
    tag_id: int | None = None,
    tag_name: str | None = None,
    shift_label: str,
    xlsx_path: str | Path | None = None,
) -> str | None:
    """
    Return the process number string for tag + shift from ShipStation Tags.xlsx.

    Prefer match by Tag ID; fall back to case-insensitive Tag Name.
    Returns None if the workbook/row/cell is missing or the shift is unknown.
    """
    path = Path(xlsx_path) if xlsx_path else DEFAULT_XLSX_PATH
    if not path.is_file():
        return None

    shift_key = _normalize_shift(shift_label)
    col = SHIFT_PROCESS_COLS.get(shift_key)
    if col is None:
        return None

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
        name_key = _name_key(tag_name or "")
        want_id = int(tag_id) if tag_id is not None else None

        by_id_row: tuple | None = None
        by_name_row: tuple | None = None

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < col:
                continue
            row_name = row[1] if len(row) > 1 else None
            row_id = _as_int(row[2] if len(row) > 2 else None)
            if want_id is not None and row_id == want_id:
                by_id_row = row
                break
            if name_key and by_name_row is None and _name_key(str(row_name or "")) == name_key:
                by_name_row = row

        chosen = by_id_row if by_id_row is not None else by_name_row
        if chosen is None:
            return None
        return _cell_to_process_str(chosen[col - 1])
    finally:
        wb.close()


def resolve_process_number(
    gui_value: str | None,
    *,
    tag_id: int | None = None,
    tag_name: str | None = None,
    shift_label: str,
    xlsx_path: str | Path | None = None,
) -> str | None:
    """
    Prefer a non-empty GUI process value; otherwise look up Tags.xlsx.

    Returns None when neither source yields a process number.
    """
    manual = (gui_value or "").strip()
    if manual:
        return manual
    return lookup_process_number(
        tag_id=tag_id,
        tag_name=tag_name,
        shift_label=shift_label,
        xlsx_path=xlsx_path,
    )


def resolve_tag_list_processes(
    tags: list[tuple[int, str]],
    *,
    shift_label: str,
    gui_value: str | None = None,
    xlsx_path: str | Path | None = None,
) -> tuple[list[tuple[int, str, str]], str | None]:
    """
    Resolve a process number for each selected tag.

    - One tag: non-empty ``gui_value`` wins; otherwise Tags.xlsx.
    - Multiple tags: ``gui_value`` is ignored; each tag must have a sheet value.
      Duplicate process numbers across tags are an error.

    Returns ``(resolved, None)`` on success where each item is
    ``(tag_id, tag_name, process_number)``, or ``([], error_message)`` on failure.
    """
    if not tags:
        return [], "Please select at least one ShipStation tag."

    shift = (shift_label or "").strip()
    path = Path(xlsx_path) if xlsx_path else DEFAULT_XLSX_PATH
    multi = len(tags) > 1
    resolved: list[tuple[int, str, str]] = []
    seen_process: dict[str, str] = {}

    for tag_id, tag_name in tags:
        if multi:
            process = lookup_process_number(
                tag_id=tag_id,
                tag_name=tag_name,
                shift_label=shift,
                xlsx_path=path,
            )
            if not process:
                return [], (
                    "Process number is required for each selected ShipStation tag.\n\n"
                    f"None found in:\n{path}\n"
                    f"for tag '{tag_name}' (id {tag_id}), shift '{shift}'.\n"
                    "Fill the Process No column for that shift, or select fewer tags."
                )
        else:
            process = resolve_process_number(
                gui_value,
                tag_id=tag_id,
                tag_name=tag_name,
                shift_label=shift,
                xlsx_path=path,
            )
            if not process:
                return [], (
                    "Process number is required when using a ShipStation tag.\n\n"
                    f"No value in the field and none found in:\n{path}\n"
                    f"for tag '{tag_name}' (id {tag_id}), shift '{shift}'.\n"
                    "Enter a process number or fill the Process No column for that shift."
                )

        other = seen_process.get(process)
        if other is not None:
            return [], (
                f"Two selected tags resolve to the same process number '{process}' "
                f"for shift '{shift}':\n"
                f"  '{other}' and '{tag_name}'\n\n"
                "Change the Process No values in the Tags sheet, or deselect one tag."
            )
        seen_process[process] = tag_name
        resolved.append((tag_id, tag_name, process))

    return resolved, None


def parse_shipstation_tags_config(data: dict) -> list[tuple[int, str]]:
    """Load selected tags from config; prefer ``shipstation_tags``, else legacy scalars."""
    raw = data.get("shipstation_tags")
    out: list[tuple[int, str]] = []
    if isinstance(raw, list):
        for item in raw:
            if not isinstance(item, dict):
                continue
            try:
                tag_id = int(item.get("id"))
            except (TypeError, ValueError):
                continue
            name = str(item.get("name") or "").strip()
            if not name:
                continue
            if any(existing_id == tag_id for existing_id, _ in out):
                continue
            out.append((tag_id, name))
        if out:
            return out

    raw_id = str(data.get("shipstation_tag_id") or "").strip()
    name = str(data.get("shipstation_tag_name") or "").strip()
    if raw_id and name:
        try:
            return [(int(raw_id), name)]
        except ValueError:
            return []
    return []


def shipstation_tags_config_payload(
    tags: list[tuple[int, str]],
) -> tuple[list[dict[str, object]], str, str]:
    """Return (shipstation_tags list, legacy name, legacy id) for config save."""
    payload = [{"id": tag_id, "name": name} for tag_id, name in tags]
    if tags:
        first_id, first_name = tags[0]
        return payload, first_name, str(first_id)
    return payload, "", ""
