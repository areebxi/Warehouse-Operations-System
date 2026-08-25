from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import Workbook

from .helpers import _base_additional_no_dash, _extended_process_and_item_number, _gender_colour_size_combo_hyphenated, _item_number_from_extended, _normalize, _order_number_base, _process_number_for_excel_from_row, _process_plus_additional, _remap_dtf_item_sku, _split_item_sku_by_lg, _tracker_seq_from_val


def _write_picking(df: pd.DataFrame, process_base: str, picking_number: str, dispatch_date: date, path: Path) -> None:
    rows_out = []
    for _, row in df.iterrows():
        qty = int(row.get("Item Quantity", 1) or 1)
        if qty < 1:
            qty = 1
        process_and_item = row.get("Process and Item Number", "")
        extended = _extended_process_and_item_number(process_and_item)
        process_excel = _process_number_for_excel_from_row(process_and_item)
        item_num = _item_number_from_extended(extended)
        for i in range(qty):
            bulk = qty - i
            rows_out.append({
                "dispatch_date": dispatch_date,
                "Picking Number": picking_number if not rows_out else "",
                "Process Number": process_excel,
                "Item Number": item_num,
                "Custom Label": _normalize(row.get("Item SKU", "")),
                "Gender-Apparel": _normalize(row.get("Gender Apparel", "")),
                "Color": _normalize(row.get("Colour", "")),
                "Size": _normalize(row.get("Size", "")),
                "Qty": qty,
                "Bulk": bulk,
                "Missing Apparel": "",
                "Status": "",
                "Order Number": _order_number_base(row),
            })
    if not rows_out:
        wb = Workbook()
        ws = wb.active
        ws.title = "Picking"
        ws.cell(row=1, column=1, value="Date")
        wb.save(path)
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Picking"
    ws.cell(row=1, column=1, value="Date")
    ws.cell(row=1, column=2, value="Picking Number")
    ws.cell(row=1, column=3, value="Process Number")
    ws.cell(row=1, column=4, value="Item Number")
    ws.cell(row=1, column=5, value="Custom Label")
    ws.cell(row=1, column=6, value="Gender-Apparel")
    ws.cell(row=1, column=7, value="Color")
    ws.cell(row=1, column=8, value="Size")
    ws.cell(row=1, column=9, value="Qty")
    ws.cell(row=1, column=10, value="Bulk")
    ws.cell(row=1, column=11, value="Missing Apparel")
    ws.cell(row=1, column=12, value="Status")
    ws.cell(row=1, column=27, value="Order Number (Base)")
    ws.cell(row=1, column=28, value="Process Number")
    ws.cell(row=1, column=29, value="Item Number")
    ws.cell(row=1, column=53, value="Process Number")
    ws.cell(row=1, column=54, value="Item Number")
    ws.cell(row=1, column=55, value="Picking Number")
    ws.cell(row=1, column=56, value="Gender-Apparel")
    ws.cell(row=1, column=57, value="Color")
    ws.cell(row=1, column=58, value="Size")
    ws.cell(row=1, column=59, value="Bulk")
    ws.cell(row=1, column=60, value="Missing Apparel")
    ws.cell(row=1, column=61, value="Status")

    date_str = dispatch_date.strftime("%d-%m-%Y")
    for r, d in enumerate(rows_out, start=2):
        pref_process = f"Process {d['Process Number']}" if d["Process Number"] else ""
        pref_item = f"Item-{d['Item Number']}" if d["Item Number"] else ""
        ws.cell(row=r, column=1, value=date_str)
        ws.cell(row=r, column=2, value=d["Picking Number"])
        ws.cell(row=r, column=3, value=d["Process Number"])
        ws.cell(row=r, column=4, value=d["Item Number"])
        ws.cell(row=r, column=5, value=d["Custom Label"])
        ws.cell(row=r, column=6, value=d["Gender-Apparel"])
        ws.cell(row=r, column=7, value=d["Color"])
        ws.cell(row=r, column=8, value=d["Size"])
        ws.cell(row=r, column=9, value=d["Qty"])
        ws.cell(row=r, column=10, value=d["Bulk"])
        ws.cell(row=r, column=11, value=d["Missing Apparel"])
        ws.cell(row=r, column=12, value=d["Status"])
        ws.cell(row=r, column=27, value=d["Order Number"])
        ws.cell(row=r, column=28, value=pref_process)
        ws.cell(row=r, column=29, value=pref_item)
        ws.cell(row=r, column=53, value=d["Process Number"])
        ws.cell(row=r, column=54, value=d["Item Number"])
        ws.cell(row=r, column=55, value=d["Picking Number"])
        ws.cell(row=r, column=56, value=d["Gender-Apparel"])
        ws.cell(row=r, column=57, value=d["Color"])
        ws.cell(row=r, column=58, value=d["Size"])
        ws.cell(row=r, column=59, value=d["Bulk"])
        ws.cell(row=r, column=60, value=d["Missing Apparel"])
        ws.cell(row=r, column=61, value=d["Status"])

    wb.save(path)


def _write_orders_details(df: pd.DataFrame, process_base: str, path: Path) -> None:
    if df.empty:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        wb.save(path)
        return

    order_counts = df["Order Number (Base)"].fillna("").astype(str).value_counts()
    df = df.copy()
    df["_extended"] = df["Process and Item Number"].map(_extended_process_and_item_number)
    df["_block"] = df["_extended"].map(_process_plus_additional)
    block_recipient_count = df.groupby("_block")["Recipient Name"].nunique().to_dict()
    df["_qty"] = df["Item Quantity"].fillna(1).astype(int)
    block_quantity = df.groupby("_block")["_qty"].sum().to_dict()

    def _row_is_merge(r: pd.Series) -> bool:
        order_num = _order_number_base(r)
        is_repeated_order = order_counts.get(order_num, 0) >= 2
        has_multi_qty = r["_qty"] > 1
        return is_repeated_order or has_multi_qty

    df["_is_merge_row"] = df.apply(_row_is_merge, axis=1)

    seen = set()
    blocks_ordered = []
    for b in df["_block"]:
        if b not in seen:
            seen.add(b)
            blocks_ordered.append(b)

    headers = [
        "Condition", "Gender Apparel-Colour Name-Size", "Merge", "Single", "Single",
        "Porcess Number", "Gender Apparel-Colour Name-Size", "H", "Type", "J", "K",
        "Condition", "Porcess Number", "N", "O", "Gender Apparel-Colour Name-Size", "Q",
        "Porcess Number", "Merge", "Single",
    ]
    data = []
    for block in blocks_ordered:
        grp = df[df["_block"] == block]
        first = grp.iloc[0]
        is_merge = bool(grp["_is_merge_row"].any())
        merge_val = block_recipient_count.get(block, 0)
        single_val = int(block_quantity.get(block, 0))

        cond = "Condition 1 Merge" if is_merge else "Condition 4"
        combo = "Merge Orders" if is_merge else _gender_colour_size_combo_hyphenated(first)
        customise = _normalize(first.get("Customise", ""))
        type_base = "Personalised" if customise.lower() == "yes" else "Normal"
        i_val = f"{type_base}-Merge" if is_merge else type_base

        process_excel = _process_number_for_excel_from_row(first.get("Process and Item Number", ""))
        data.append([
            cond, combo, merge_val, single_val, single_val,
            process_excel, combo, "", i_val, "", "", cond, process_excel, "", "", combo, "",
            process_excel, merge_val, single_val,
        ])

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    for r, row_data in enumerate(data, start=2):
        for c, val in enumerate(row_data, start=1):
            ws.cell(row=r, column=c, value=val)
    wb.save(path)


def _write_dtf_des(
    df: pd.DataFrame,
    path: Path,
    use_fixed_process_number: bool = False,
    use_fixed_numeric_process: bool = False,
    dtf_sku_map: dict[str, str] | None = None,
) -> None:
    headers = [
        "Order - Number", "Item - Qty", "Item - SKU", "Item - Name", "Ship To - Name",
        "Notes - From Buyer", "Ship To - Postal Code", "Source", "Process Num", "Genre",
        "Order Type", "Orders Type Abbrevation", "Condition", "Customise", "", "Item Num",
    ]
    if df.empty:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        for c, h in enumerate(headers, start=1):
            ws.cell(row=1, column=c, value=h)
        wb.save(path)
        return

    rows = []
    for _, row in df.iterrows():
        process_and_item = row.get("Process and Item Number", "")
        extended = _extended_process_and_item_number(process_and_item)
        if use_fixed_numeric_process:
            seq = _tracker_seq_from_val(process_and_item)
            if seq is not None:
                process_excel = seq
            else:
                process_excel = _process_number_for_excel_from_row(process_and_item)
        elif use_fixed_process_number:
            process_excel = _base_additional_no_dash(extended)
        else:
            process_excel = _process_number_for_excel_from_row(process_and_item)
        process_display = f"Process {process_excel}" if process_excel else ""
        item_num_str = _item_number_from_extended(extended)
        item_display = f"Item {item_num_str}" if item_num_str else "Item 1"
        item_sku = _normalize(row.get("Item SKU", ""))
        customise_val = _normalize(row.get("Customise", ""))
        is_customised = customise_val.lower() == "yes"
        segments = [item_sku] if is_customised else _split_item_sku_by_lg(item_sku)
        m = dtf_sku_map or {}
        segments = [_remap_dtf_item_sku(seg, m) for seg in segments]
        qty = max(1, int(row.get("Item Quantity", 1) or 1))
        base_row = [
            _order_number_base(row),
            1,
            item_sku,
            _normalize(row.get("Item Name", "")),
            _normalize(row.get("Recipient Name", "")),
            "", "", "", process_display, _normalize(row.get("Gender Apparel", "")),
            "", "", "",
        ]
        for seg in segments:
            for _ in range(qty):
                row_data = base_row.copy()
                row_data[2] = seg
                row_data.extend([customise_val, "", item_display])
                rows.append(row_data)
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    for r, row_data in enumerate(rows, start=2):
        for c, val in enumerate(row_data, start=1):
            ws.cell(row=r, column=c, value=val)
    wb.save(path)

