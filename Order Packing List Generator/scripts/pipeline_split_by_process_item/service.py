import sys
from datetime import date, datetime
from functools import partial
from pathlib import Path
from typing import Callable, Optional

import pandas as pd
from openpyxl import load_workbook

from scripts.pipeline_generate_packing_list_pdf.runtime_api import load_position_code_to_draw
from scripts.pipeline_generate_packing_list_pdf.runtime_config import DEFAULT_POSITION_CODE

from scripts.pipeline_runtime.order_number_csv import (
    coerce_order_number_columns,
    read_csv_with_order_numbers,
)

from .common import (
    _normalize,
    _normalize_key,
    _order_number_column,
    _parse_ship_by,
    _position_after_merge,
    _reorder_columns_for_output,
    _resolve_workbook_path,
    sanitize_filename,
)
from .config import (
    DEFAULT_OUTPUT_DIR,
    DEFAULT_WORKBOOK,
    PROCESS_TRACKER_SHEET,
    REQUIRED_COLUMNS,
    TRACKER_SEQUENCE_START,
)
from .grouping import _expand_df_by_quantity, _sort_and_assign_merge_first
from .size_sequence import load_sequence_by_size


def _emit(msg: str, log: Optional[Callable[[str], None]], *, err: bool = False) -> None:
    if log:
        log(msg)
    elif err:
        print(msg, file=sys.stderr)
    else:
        print(msg)


def run(
    step5_csv_path: Path,
    output_dir: Path,
    workbook_path: Path | None,
    run_date: date | None = None,
    use_simple_process_format: bool = False,
    use_fixed_numeric_process: bool = False,
    fixed_process_number: str | None = None,
    log: Optional[Callable[[str], None]] = None,
) -> None:
    df = read_csv_with_order_numbers(step5_csv_path)
    _emit(f"Step 6 (split by process): read {len(df)} rows from {step5_csv_path.name}", log)
    new_cols = []
    for c in df.columns:
        c2 = str(c).strip().lstrip("\ufeff")
        if _normalize_key(c2) == "order number" and c2 != "Order Number":
            c2 = "Order Number"
        new_cols.append(c2)
    df.columns = new_cols
    df = coerce_order_number_columns(df)
    missing = [c for c in REQUIRED_COLUMNS if c not in df.columns]
    if missing:
        raise ValueError(f"Step-5 CSV is missing required column(s): {', '.join(missing)}")

    df["_orig_idx"] = range(len(df))
    df = _expand_df_by_quantity(df)

    size_to_rank = None
    position_code_to_draw: dict[str, str] = {}
    if workbook_path is not None:
        size_to_rank = load_sequence_by_size(workbook_path)
        try:
            position_code_to_draw = load_position_code_to_draw(workbook_path)
        except Exception:
            position_code_to_draw = {}
    if size_to_rank is None:
        _emit(
            "Warning: Size sequence not loaded (workbook missing or no 'Sequence by Size' column in Process Info Sheet). Rows will not be sorted by size.",
            log,
            err=True,
        )
    else:
        _emit(
            f"Sorting by size using {len(size_to_rank)} sizes from 'Sequence by Size'.",
            log,
            err=True,
        )

    today = run_date or date.today()
    output_dir.mkdir(parents=True, exist_ok=True)

    pin = df["Process and Item Number"].fillna("").astype(str).str.strip()
    df_grouped = df.assign(_pin_key=pin)
    groups: list[tuple[str, pd.DataFrame]] = []
    for key, group in df_grouped.groupby("_pin_key", sort=False):
        groups.append((str(key), group.drop(columns=["_pin_key"])))

    group_meta: list[dict] = []
    for base, group in groups:
        process_name = sanitize_filename(base if base else "")
        is_prime = False
        if "Prime" in group.columns:
            is_prime = any(_normalize_key(v) == "yes" for v in group["Prime"])
        is_dispatch_today = False
        if "Ship By" in group.columns:
            for v in group["Ship By"]:
                ship_date = _parse_ship_by(v)
                if ship_date is not None and ship_date == today:
                    is_dispatch_today = True
                    break
        group_meta.append(
            {
                "base": base,
                "process_name": process_name,
                "is_prime": is_prime,
                "is_dispatch_today": is_dispatch_today,
            }
        )

    base_to_sequence: dict[str, int] = {}

    tracker_path: Path | None = None
    if use_simple_process_format or use_fixed_numeric_process:
        tracker_path = None
    elif workbook_path is not None:
        try:
            tracker_path = _resolve_workbook_path(workbook_path)
            _emit(
                f"Process Number Tracker: using workbook '{tracker_path.resolve()}'.",
                log,
                err=True,
            )
        except (FileNotFoundError, OSError) as exc:
            tracker_path = None
            _emit(
                f"Process Number Tracker: workbook not found ({exc}); skipping tracker updates.",
                log,
                err=True,
            )

    if tracker_path is not None and group_meta:
        try:
            wb = load_workbook(tracker_path)
        except Exception as exc:
            wb = None
            _emit(
                f"Process Number Tracker: failed to load workbook '{tracker_path}': {exc}",
                log,
                err=True,
            )

        if wb is not None:
            if PROCESS_TRACKER_SHEET in wb.sheetnames:
                ws = wb[PROCESS_TRACKER_SHEET]
            else:
                ws = wb.create_sheet(PROCESS_TRACKER_SHEET)
                ws.append(["Date", "Process Number", "Sequence Number"])

            existing_today: dict[str, int] = {}
            max_today_seq = 0
            last_data_row = 1
            for row in ws.iter_rows(min_row=2):
                if len(row) < 3:
                    continue
                date_cell, proc_cell, seq_cell = row[0], row[1], row[2]
                if any(c.value not in (None, "") for c in (date_cell, proc_cell, seq_cell)):
                    if date_cell.row > last_data_row:
                        last_data_row = date_cell.row

                val = date_cell.value
                if val is None:
                    continue
                cell_date: date | None = None
                if isinstance(val, datetime):
                    cell_date = val.date()
                elif isinstance(val, date):
                    cell_date = val
                elif isinstance(val, str):
                    cell_date = _parse_ship_by(val)
                if cell_date != today:
                    continue

                proc_val = _normalize(proc_cell.value)
                if not proc_val:
                    continue
                seq_val = seq_cell.value
                if seq_val is None:
                    continue
                try:
                    seq_int = int(seq_val)
                except (TypeError, ValueError):
                    continue

                existing_today[proc_val] = seq_int
                if seq_int > max_today_seq:
                    max_today_seq = seq_int

            def _tier(meta: dict) -> int:
                if meta["is_prime"]:
                    return 0
                if meta["is_dispatch_today"]:
                    return 1
                return 2

            sorted_meta = sorted(
                group_meta,
                key=lambda m: (_tier(m), m["process_name"].lower()),
            )

            next_row = last_data_row + 1 if last_data_row >= 2 else 2
            next_seq = max(max_today_seq + 1, TRACKER_SEQUENCE_START)
            new_rows_for_today = 0
            for meta in sorted_meta:
                proc_name = _normalize(meta["process_name"])
                if proc_name in existing_today:
                    base_to_sequence[proc_name] = existing_today[proc_name]
                    continue

                seq = next_seq
                next_seq += 1
                existing_today[proc_name] = seq
                base_to_sequence[proc_name] = seq
                ws.cell(row=next_row, column=1, value=today)
                ws.cell(row=next_row, column=2, value=proc_name)
                ws.cell(row=next_row, column=3, value=seq)
                next_row += 1
                new_rows_for_today += 1

            if new_rows_for_today > 0:
                try:
                    wb.save(tracker_path)
                    _emit(
                        f"Process Number Tracker: updated {new_rows_for_today} process(es) for {today}. Saved to '{tracker_path.resolve()}'. Open this file to see the sheet.",
                        log,
                        err=True,
                    )
                except Exception as exc:
                    _emit(
                        f"Process Number Tracker: failed to save to '{tracker_path}': {exc}",
                        log,
                        err=True,
                    )

    written = 0
    for base, group in groups:
        process_name = sanitize_filename(base if base else "")
        seq = base_to_sequence.get(_normalize(process_name))
        group_sorted = _sort_and_assign_merge_first(
            group,
            size_to_rank,
            sequence_number=seq,
            use_simple_process_format=use_simple_process_format,
            use_fixed_numeric_process=use_fixed_numeric_process,
            fixed_process_number=fixed_process_number,
        )
        if (
            "Position" in group_sorted.columns
            and "Logo/Design Image" in group_sorted.columns
            and "Item SKU" in group_sorted.columns
        ):
            group_sorted = group_sorted.copy()
            position_merge = partial(
                _position_after_merge,
                position_code_to_draw=position_code_to_draw or None,
                default_position_code=DEFAULT_POSITION_CODE,
            )
            group_sorted["Position"] = group_sorted.apply(position_merge, axis=1)
        path = output_dir / f"{process_name}.csv"
        group_sorted = group_sorted.drop(columns=["_orig_idx"], errors="ignore")
        group_sorted = _reorder_columns_for_output(group_sorted)
        group_sorted.to_csv(path, index=False, encoding="utf-8")
        written += 1

    if log and written:
        proc_names = sorted({sanitize_filename(base if base else "") for base, _ in groups})
        for n in proc_names[:35]:
            _emit(f"  {n}.csv", log)
        if len(proc_names) > 35:
            _emit(f"  ... and {len(proc_names) - 35} more process file(s)", log)

    _emit(f"Split into {written} process CSV file(s) under {output_dir.resolve()}", log)


__all__ = ["run", "DEFAULT_OUTPUT_DIR", "DEFAULT_WORKBOOK"]

