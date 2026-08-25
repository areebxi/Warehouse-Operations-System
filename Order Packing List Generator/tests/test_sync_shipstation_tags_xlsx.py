"""Unit tests for ShipStation Tags.xlsx sync helper."""

from pathlib import Path

from openpyxl import Workbook

from scripts.pipeline_shipstation.sync_tags_xlsx import (
    HEADERS,
    SHEET_NAME,
    sync_shipstation_tags_xlsx,
)


class _FakeClient:
    def __init__(self, tags: list[dict]) -> None:
        self._tags = tags

    def list_tags(self) -> list[dict]:
        return list(self._tags)


def _write_workbook(path: Path, rows: list[tuple[int, str, int | None]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    for col, header in enumerate(HEADERS, start=1):
        ws.cell(1, col).value = header
    for i, (sr, name, tid) in enumerate(rows, start=2):
        ws.cell(i, 1).value = sr
        ws.cell(i, 2).value = name
        ws.cell(i, 3).value = tid
        ws.cell(i, 4).value = 1000 + i  # process col should be preserved
    wb.save(path)


def test_sync_adds_and_updates_ids(tmp_path: Path):
    xlsx = tmp_path / "ShipStation Tags.xlsx"
    _write_workbook(
        xlsx,
        [
            (1, "Alpha", 10),
            (2, "Beta", 20),
        ],
    )
    client = _FakeClient(
        [
            {"tagId": 11, "name": "Alpha"},  # id changed
            {"tagId": 20, "name": "Beta"},
            {"tagId": 30, "name": "Gamma"},  # new
        ]
    )

    result = sync_shipstation_tags_xlsx(xlsx, client=client, backup=False)

    assert len(result.updated_ids) == 1
    assert result.updated_ids[0] == ("Alpha", 10, 11)
    assert result.added == [("Gamma", 30)]
    assert result.obsolete == []

    from openpyxl import load_workbook

    wb = load_workbook(xlsx)
    ws = wb[SHEET_NAME]
    # Alpha id updated
    assert ws.cell(2, 3).value == 11
    # Beta process column preserved
    assert ws.cell(3, 4).value == 1003
    # Gamma appended
    assert ws.cell(4, 2).value == "Gamma"
    assert ws.cell(4, 3).value == 30
    assert ws.cell(4, 1).value == 3


def test_sync_dry_run_does_not_write(tmp_path: Path):
    xlsx = tmp_path / "ShipStation Tags.xlsx"
    _write_workbook(xlsx, [(1, "Alpha", 10)])
    mtime_before = xlsx.stat().st_mtime_ns
    client = _FakeClient([{"tagId": 11, "name": "Alpha"}, {"tagId": 30, "name": "Gamma"}])

    result = sync_shipstation_tags_xlsx(xlsx, client=client, dry_run=True)

    assert result.dry_run is True
    assert len(result.added) == 1
    assert len(result.updated_ids) == 1
    assert xlsx.stat().st_mtime_ns == mtime_before


def test_sync_reports_obsolete(tmp_path: Path):
    xlsx = tmp_path / "ShipStation Tags.xlsx"
    _write_workbook(xlsx, [(1, "OldTag", 99), (2, "Keep", 1)])
    client = _FakeClient([{"tagId": 1, "name": "Keep"}])

    result = sync_shipstation_tags_xlsx(xlsx, client=client, backup=False)

    assert result.obsolete == [("OldTag", 99)]
    assert result.added == []
