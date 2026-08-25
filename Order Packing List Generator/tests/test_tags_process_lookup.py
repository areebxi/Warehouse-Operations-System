"""Unit tests for ShipStation Tags.xlsx process-number lookup."""

from pathlib import Path

from openpyxl import Workbook

from scripts.pipeline_shipstation.sync_tags_xlsx import HEADERS, SHEET_NAME
from scripts.pipeline_shipstation.tags_process_lookup import (
    lookup_process_number,
    parse_shipstation_tags_config,
    resolve_process_number,
    resolve_tag_list_processes,
    shipstation_tags_config_payload,
)


def _write_tags_workbook(
    path: Path,
    rows: list[tuple[int, str, int | None, object, object, object, object, object]],
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    for col, header in enumerate(HEADERS, start=1):
        ws.cell(1, col).value = header
    for i, row in enumerate(rows, start=2):
        for col, value in enumerate(row, start=1):
            ws.cell(i, col).value = value
    wb.save(path)


def test_lookup_by_tag_id(tmp_path: Path):
    xlsx = tmp_path / "ShipStation Tags.xlsx"
    _write_tags_workbook(
        xlsx,
        [
            (1, "Alpha", 10, 8000, 8100, None, None, None),
            (2, "Beta", 20, 9000, None, None, None, None),
        ],
    )
    assert (
        lookup_process_number(tag_id=10, tag_name="Wrong", shift_label="1st", xlsx_path=xlsx)
        == "8000"
    )
    assert (
        lookup_process_number(tag_id=10, tag_name="Alpha", shift_label="2nd", xlsx_path=xlsx)
        == "8100"
    )


def test_lookup_by_tag_name_case_insensitive(tmp_path: Path):
    xlsx = tmp_path / "ShipStation Tags.xlsx"
    _write_tags_workbook(
        xlsx,
        [
            (1, "Alpha Tag", 10, 8000, 8100, None, None, None),
        ],
    )
    assert (
        lookup_process_number(tag_id=999, tag_name="alpha tag", shift_label="1st", xlsx_path=xlsx)
        == "8000"
    )


def test_lookup_blank_cell_returns_none(tmp_path: Path):
    xlsx = tmp_path / "ShipStation Tags.xlsx"
    _write_tags_workbook(
        xlsx,
        [
            (1, "Alpha", 10, 8000, None, None, None, None),
        ],
    )
    assert (
        lookup_process_number(tag_id=10, tag_name="Alpha", shift_label="2nd", xlsx_path=xlsx)
        is None
    )


def test_lookup_missing_workbook_returns_none(tmp_path: Path):
    missing = tmp_path / "missing.xlsx"
    assert (
        lookup_process_number(tag_id=10, tag_name="Alpha", shift_label="1st", xlsx_path=missing)
        is None
    )


def test_lookup_normalizes_float_process(tmp_path: Path):
    xlsx = tmp_path / "ShipStation Tags.xlsx"
    _write_tags_workbook(
        xlsx,
        [
            (1, "Alpha", 10, 8000.0, None, None, None, None),
        ],
    )
    assert (
        lookup_process_number(tag_id=10, tag_name="Alpha", shift_label="1st", xlsx_path=xlsx)
        == "8000"
    )


def test_lookup_accepts_shift_with_suffix(tmp_path: Path):
    xlsx = tmp_path / "ShipStation Tags.xlsx"
    _write_tags_workbook(
        xlsx,
        [
            (1, "Alpha", 10, 8000, 8100, None, None, None),
        ],
    )
    assert (
        lookup_process_number(tag_id=10, tag_name="Alpha", shift_label="2nd Shift", xlsx_path=xlsx)
        == "8100"
    )


def test_resolve_gui_value_wins(tmp_path: Path):
    xlsx = tmp_path / "ShipStation Tags.xlsx"
    _write_tags_workbook(
        xlsx,
        [
            (1, "Alpha", 10, 8000, None, None, None, None),
        ],
    )
    assert (
        resolve_process_number(
            "1234",
            tag_id=10,
            tag_name="Alpha",
            shift_label="1st",
            xlsx_path=xlsx,
        )
        == "1234"
    )
    assert (
        resolve_process_number(
            "  ",
            tag_id=10,
            tag_name="Alpha",
            shift_label="1st",
            xlsx_path=xlsx,
        )
        == "8000"
    )
    assert (
        resolve_process_number(
            "",
            tag_id=10,
            tag_name="Alpha",
            shift_label="2nd",
            xlsx_path=xlsx,
        )
        is None
    )


def test_resolve_tag_list_multi_uses_sheet_and_ignores_gui(tmp_path: Path):
    xlsx = tmp_path / "ShipStation Tags.xlsx"
    _write_tags_workbook(
        xlsx,
        [
            (1, "Alpha", 10, 8000, None, None, None, None),
            (2, "Beta", 20, 9000, None, None, None, None),
        ],
    )
    resolved, err = resolve_tag_list_processes(
        [(10, "Alpha"), (20, "Beta")],
        shift_label="1st",
        gui_value="1111",
        xlsx_path=xlsx,
    )
    assert err is None
    assert resolved == [(10, "Alpha", "8000"), (20, "Beta", "9000")]


def test_resolve_tag_list_duplicate_process_errors(tmp_path: Path):
    xlsx = tmp_path / "ShipStation Tags.xlsx"
    _write_tags_workbook(
        xlsx,
        [
            (1, "Alpha", 10, 8000, None, None, None, None),
            (2, "Beta", 20, 8000, None, None, None, None),
        ],
    )
    resolved, err = resolve_tag_list_processes(
        [(10, "Alpha"), (20, "Beta")],
        shift_label="1st",
        xlsx_path=xlsx,
    )
    assert resolved == []
    assert err is not None
    assert "same process number" in err


def test_resolve_tag_list_single_gui_wins(tmp_path: Path):
    xlsx = tmp_path / "ShipStation Tags.xlsx"
    _write_tags_workbook(
        xlsx,
        [
            (1, "Alpha", 10, 8000, None, None, None, None),
        ],
    )
    resolved, err = resolve_tag_list_processes(
        [(10, "Alpha")],
        shift_label="1st",
        gui_value="1234",
        xlsx_path=xlsx,
    )
    assert err is None
    assert resolved == [(10, "Alpha", "1234")]


def test_parse_shipstation_tags_config_prefers_list_and_migrates_legacy():
    assert parse_shipstation_tags_config(
        {
            "shipstation_tags": [{"id": 1, "name": "A"}, {"id": 2, "name": "B"}],
            "shipstation_tag_id": "99",
            "shipstation_tag_name": "Legacy",
        }
    ) == [(1, "A"), (2, "B")]
    assert parse_shipstation_tags_config(
        {"shipstation_tag_id": "99", "shipstation_tag_name": "Legacy"}
    ) == [(99, "Legacy")]
    assert shipstation_tags_config_payload([(1, "A"), (2, "B")]) == (
        [{"id": 1, "name": "A"}, {"id": 2, "name": "B"}],
        "A",
        "1",
    )
