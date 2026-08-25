"""Tests for DTF Des Customise column (Step 7)."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from scripts.pipeline_generate_excel_outputs.writers import _write_dtf_des


def _minimal_row(**overrides) -> dict:
    row = {
        "Order Number (Base)": "ORD-001",
        "Item Quantity": 1,
        "Item SKU": "123456LG-M-T-BLK-M",
        "Item Name": "Test Item",
        "Recipient Name": "Jane Doe",
        "Process and Item Number": "Process 100 Item-1",
        "Gender Apparel": "Mens-T-Shirt",
        "Customise": "",
    }
    row.update(overrides)
    return row


def test_dtf_des_customise_column_headers_and_values(tmp_path: Path):
    df = pd.DataFrame(
        [
            _minimal_row(
                **{
                    "Customise": "Yes",
                    "Item SKU": "98765PER-M-T-BLK-2XL",
                    "Process and Item Number": "Process 100 Item-1",
                }
            ),
            _minimal_row(
                **{
                    "Customise": "",
                    "Item SKU": "123456LG-M-T-BLK-M",
                    "Process and Item Number": "Process 100 Item-2",
                }
            ),
        ]
    )
    out = tmp_path / "DTF Des-P100.xlsx"
    _write_dtf_des(df, out)

    wb = load_workbook(out)
    ws = wb.active
    assert ws.cell(row=1, column=14).value == "Customise"
    assert ws.cell(row=1, column=16).value == "Item Num"

    assert ws.cell(row=2, column=14).value == "Yes"
    assert ws.cell(row=2, column=3).value == "98765PER-M-T-BLK-2XL"

    assert ws.cell(row=3, column=14).value in ("", None)
    assert ws.cell(row=3, column=3).value == "123456LG-M-T-BLK-M"


def test_dtf_des_custom_row_sku_not_split(tmp_path: Path):
    df = pd.DataFrame(
        [
            _minimal_row(
                **{
                    "Customise": "Yes",
                    "Item SKU": "111LG-M-T-BLK-M-222LG-M-T-WHI-L",
                    "Process and Item Number": "Process 100 Item-1",
                }
            ),
        ]
    )
    out = tmp_path / "DTF Des-P100.xlsx"
    _write_dtf_des(df, out)

    wb = load_workbook(out)
    ws = wb.active
    assert ws.max_row == 2
    assert ws.cell(row=2, column=3).value == "111LG-M-T-BLK-M-222LG-M-T-WHI-L"
    assert ws.cell(row=2, column=14).value == "Yes"


def test_dtf_des_empty_dataframe_still_has_customise_header(tmp_path: Path):
    out = tmp_path / "DTF Des-P100.xlsx"
    _write_dtf_des(pd.DataFrame(), out)

    wb = load_workbook(out)
    ws = wb.active
    assert ws.cell(row=1, column=14).value == "Customise"
    assert ws.cell(row=1, column=16).value == "Item Num"
