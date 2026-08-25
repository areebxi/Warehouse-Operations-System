"""
PDF Generator - Extracted from plain-orders cursor ai.py
This file contains only the PDF generation functionality for packing slips.
"""

import pandas as pd
from fpdf import FPDF
from pathlib import Path
from datetime import date

from app_paths import APP_ROOT, asset_path, data_path, packs_database_path, product_database_path, tag_output_dir

# --- PDF Layout Constants ---
PAGE_WIDTH = 297
PAGE_HEIGHT = 210
MARGIN = 10
PRODUCT_IMG_X = MARGIN
PRODUCT_IMG_Y = 45
PRODUCT_IMG_W = 75
PRODUCT_IMG_H = 95.55
DETAILS_X_START = PRODUCT_IMG_X + PRODUCT_IMG_W + 10
BRAND_LOGO_W = 40
BRAND_LOGO_X = PAGE_WIDTH - MARGIN - BRAND_LOGO_W

# --- Column Names ---
COLUMN_NAMES = {
    # In orders.csv (PLAIN_ITEMS_CSV)
    "order_id": "Order",
    "sku": "Item SKU",
    "recipient": "Recipient",
    "quantity": "Quantity",
    "process": "Tag",
    "components": "Components",
    "component_colours": "Component Colours",

    # In Database.xlsx (PRODUCT_DATABASE_FILE)
    "product_code": "Product Code",
    "db_sku": "SKU",
    "brand": "Brand",
    "colour": "Colour",
    "size": "Size",
    "description": "Description",
    "package": "Package",
    "product_image_filename": "Product_Image_URL",
    "brand_image_filename": "Brand_Image_URL",
}

# --- File Paths ---
PLAIN_ITEMS_CSV = "packing_list_tag_30885_20250908_130713.csv"  # Example only (CLI mode)
PRODUCT_DATABASE_FILE = product_database_path()
PRODUCT_EXPORT_FILE = data_path("ProductExport.csv")
PACKS_DATABASE_FILE = packs_database_path()

SCRIPT_DIR = APP_ROOT
PRODUCT_IMAGE_FOLDER = asset_path("product_images")
BRAND_IMAGE_FOLDER = asset_path("brand_logos")

class PDF(FPDF):
    """Generates one packing slip page per item."""

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        # Force built-in Helvetica (ASCII) to avoid any machine-specific font paths.
        self.base_font = 'Helvetica'
        self._supports_unicode = False
        self.set_font(self.base_font, '', 12)

    def _t(self, value: object) -> str:
        """Return text safe for current font. Converts to latin-1 if unicode not supported."""
        s = str(value) if value is not None else ''
        if self._supports_unicode:
            return s
        # Fallback: replace common symbols and strip unsupported chars
        replacements = {
            '™': 'TM', '®': 'R', '©': 'C', '–': '-', '—': '-', '’': "'", '‘': "'", '“': '"', '”': '"', '•': '-',
        }
        for k, v in replacements.items():
            s = s.replace(k, v)
        try:
            return s.encode('latin-1', 'ignore').decode('latin-1')
        except Exception:
            return s

    def _place_image(self, image_path, x, y, w=0, h=0):
        """Helper method to check for and place an image, handling errors."""
        if not isinstance(image_path, str) or not image_path.strip():
            return

        path_obj = Path(image_path)
        if not path_obj.exists():
            print(f"      Warning: Image file not found at: {path_obj}")
            return

        # Internal helper to call fpdf.image safely even if Pillow is unavailable in runtime (e.g., EXE)
        def _safe_fpfd_image(path_str, x_val, y_val, w_val=0, h_val=0):
            try:
                self.image(path_str, x=x_val, y=y_val, w=w_val, h=h_val)
            except Exception as img_err:
                print(f"      Warning: Could not place image '{path_obj}': {img_err}. Skipping image.")

        # If both width and height are specified, maintain aspect ratio
        if w > 0 and h > 0:
            try:
                from PIL import Image
                with Image.open(path_obj) as img:
                    img_ratio = img.size[0] / img.size[1]  # width/height
                    box_ratio = w / h
                    
                    if img_ratio > box_ratio:
                        # Image is wider than box, fit to width
                        new_w = w
                        new_h = w / img_ratio
                    else:
                        # Image is taller than box, fit to height
                        new_h = h
                        new_w = h * img_ratio
                    
                    # Center the image in the box
                    offset_x = (w - new_w) / 2
                    offset_y = (h - new_h) / 2
                    
                    _safe_fpfd_image(str(path_obj), x + offset_x, y + offset_y, new_w, new_h)
            except ImportError:
                # Fallback to original behavior if PIL not available
                _safe_fpfd_image(str(path_obj), x, y, w, h)
            except Exception as e:
                print(f"      Warning: Error processing image {path_obj}: {e}")
                # Fallback to original behavior
                _safe_fpfd_image(str(path_obj), x, y, w, h)
        else:
            _safe_fpfd_image(str(path_obj), x, y, w, h)

    def _draw_header(self, item_data, product_data, item_count, total_items):
        # Use Process No from CSV if available, otherwise fallback to Tag ID
        process_value = item_data.get('Process No', '') or item_data.get(COLUMN_NAMES['process'], '')
        self.set_font(self.base_font, 'B', 24)
        self.cell(w=100, h=10, txt=self._t(f"Process: {process_value}"))

        self.set_font(self.base_font, 'B', 12)
        item_counter_text = f"Item {item_count} of {total_items}" if total_items > 1 else ""
        self.cell(w=90, h=10, txt=self._t(item_counter_text), align='C')

        brand_filename = product_data.get(COLUMN_NAMES['brand_image_filename']) if isinstance(product_data, dict) else None
        logo_path = BRAND_IMAGE_FOLDER / str(brand_filename) if brand_filename else None
        if logo_path:
            self._place_image(str(logo_path), x=BRAND_LOGO_X, y=MARGIN, w=BRAND_LOGO_W)

        self.ln(10)
        self.set_font(self.base_font, 'B', 24)
        self.set_x(90)
        self.cell(w=110, h=12, txt=self._t(item_data.get(COLUMN_NAMES['order_id'], 'N/A')), align='C')
        self.ln(15)

    def _draw_product_details(self, item_data, product_data, total_items, pack_title=None):
        _colour_img_by_uid = _load_colour_image_basenames_by_uid()

        def _find_product_image(product_filename, sku_value):
            img_path, img_source, _ = _resolve_product_image_path(
                product_filename, sku_value, colour_img_by_uid=_colour_img_by_uid
            )
            if img_path:
                print(f"      Image: using {img_source} -> {img_path}")
                return img_path
            return None

        product_filename = (product_data or {}).get(COLUMN_NAMES['product_image_filename'])
        fallback_sku = item_data.get(COLUMN_NAMES['sku'])
        image_path = _find_product_image(product_filename, fallback_sku)
        if image_path:
            self._place_image(str(image_path), x=PRODUCT_IMG_X, y=PRODUCT_IMG_Y, w=PRODUCT_IMG_W, h=PRODUCT_IMG_H)

        name_y_pos = PRODUCT_IMG_Y + PRODUCT_IMG_H + 3
        self.set_xy(PRODUCT_IMG_X, name_y_pos)
        # Slightly smaller font so two lines can fit if needed
        self.set_font(self.base_font, 'B', 20)
        self.multi_cell(w=PRODUCT_IMG_W, h=8, txt=self._t(item_data.get(COLUMN_NAMES['recipient'], 'N/A')), border=0, align='C')

        self.set_xy(DETAILS_X_START, PRODUCT_IMG_Y)

        def write_detail(label, value):
            self.set_font(self.base_font, 'B', 12)
            self.set_x(DETAILS_X_START)
            self.cell(w=40, h=8, txt=self._t(label), border=0)
            self.set_font(self.base_font, '', 12)
            self.multi_cell(w=0, h=8, txt=self._t(f": {value}"), border=0)

        # Use pack_title from Packs Database.xlsx Column AI if available, otherwise fallback to Description
        title_value = pack_title if pack_title and str(pack_title).strip() else str((product_data or {}).get(COLUMN_NAMES['description'], 'N/A'))
        write_detail('Title', title_value)
        write_detail('Colour', str((product_data or {}).get(COLUMN_NAMES['colour'], 'N/A')))
        write_detail('Size', str((product_data or {}).get(COLUMN_NAMES['size'], 'N/A')))
        
        # Quantity styled like "Merge Order" pill if greater than 1
        quantity_value = str(item_data.get(COLUMN_NAMES['quantity'], 'N/A'))
        try:
            qty_num = int(quantity_value) if quantity_value.isdigit() else 0
            if qty_num > 1:
                self.set_x(DETAILS_X_START)
                pill_text = self._t(f"Quantity: {quantity_value}")
                self.set_font(self.base_font, 'B', 14)
                # Compute pill width with padding
                try:
                    text_w = self.get_string_width(pill_text)
                except Exception:
                    text_w = 60
                pill_w = min(120, max(60, text_w + 12))
                self.set_fill_color(255, 0, 0)
                self.set_text_color(255, 255, 255)
                self.cell(w=pill_w, h=8, txt=pill_text, fill=True, align='C')
                self.set_text_color(0, 0, 0)
                self.ln(0)
            else:
                write_detail('Quantity', quantity_value)
        except Exception:
            write_detail('Quantity', quantity_value)

        self.ln(15)

        write_detail('Our SKU', str(item_data.get(COLUMN_NAMES['sku'], 'N/A')))
        write_detail('Brand', str((product_data or {}).get(COLUMN_NAMES['brand'], 'N/A')))
        write_detail('Product Code', str((product_data or {}).get(COLUMN_NAMES['product_code'], 'N/A')))
        write_detail('Package', str((product_data or {}).get(COLUMN_NAMES['package'], 'N/A')))

        if total_items > 1:
            self.ln(5)
            self.set_x(DETAILS_X_START)
            self.set_font(self.base_font, 'B', 14)
            self.set_fill_color(255, 0, 0)
            self.set_text_color(255, 255, 255)
            self.cell(w=45, h=8, txt=self._t("Merge Order"), fill=True, align='C')
            self.set_text_color(0, 0, 0)

    def add_packing_slip(self, item_data, product_data, item_count, total_items, pack_product=None, pack_name=None, pack_title=None):
        self.add_page(orientation='L')
        self._draw_header(item_data, product_data, item_count, total_items)
        self._draw_product_details(item_data, product_data, total_items, pack_title)

    def add_pack_slip(self, item_data, component_products, item_count, total_items, pack_product=None, pack_name=None, pack_title=None):
        """Render Pack of N in the same layout as single item: composite thumbnails inside the main image box, and full details on the right.

        pack_product: details row for the PACK SKU itself (used for Pack Name, etc.)
        """
        self.add_page(orientation='L')

        # Use first component as representative for header/details
        main_product = component_products[0] if component_products else {}
        # Draw standard header (process + order id + brand logo)
        self._draw_header(item_data, main_product, item_count, total_items)

        # Draw composite thumbnails within the image box region
        box_x = PRODUCT_IMG_X
        box_y = PRODUCT_IMG_Y
        box_w = PRODUCT_IMG_W
        # Make the thumbnail area taller for packs so images appear larger
        pack_size = len(component_products)
        if pack_size >= 5:
            box_h = PRODUCT_IMG_H + 40
        elif pack_size >= 3:
            box_h = PRODUCT_IMG_H + 20
        else:
            box_h = PRODUCT_IMG_H

        pack_size = len(component_products)
        cols = 2 if pack_size > 2 else pack_size  # 1xN for 1-2, 2xN for more
        rows = (pack_size + cols - 1) // cols if cols else 1
        gap = 3
        thumb_w = (box_w - (cols - 1) * gap) / max(1, cols)
        # Reserve space below each thumbnail for label: Colour, Size, Qty
        label_h = 8
        label_gap = 2
        thumb_h = ((box_h - (rows - 1) * gap) / max(1, rows)) - (label_h + label_gap)

        # Parse component colours from CSV if available
        colours_raw = item_data.get(COLUMN_NAMES['component_colours'], '')
        component_colours = []
        if isinstance(colours_raw, str) and colours_raw.strip():
            component_colours = [c.strip() for c in colours_raw.split(',') if c.strip()]
        
        _colour_img_by_uid = _load_colour_image_basenames_by_uid()

        def _find_component_image(product_filename, sku_value):
            return _resolve_product_image_path(
                product_filename, sku_value, colour_img_by_uid=_colour_img_by_uid
            )

        for idx, comp in enumerate(component_products):
            r = idx // cols
            c = idx % cols
            x = box_x + c * (thumb_w + gap)
            # Include label height and gap in row spacing so labels are not overlapped by next row images
            y = box_y + r * (thumb_h + label_h + label_gap + gap)
            img_filename = comp.get(COLUMN_NAMES['product_image_filename'])
            img_sku = comp.get(COLUMN_NAMES['db_sku'])
            img_path, _, _ = _find_component_image(img_filename, img_sku)
            if img_path:
                self._place_image(str(img_path), x=x, y=y, w=thumb_w, h=thumb_h)
            
            # Draw label under the image: Colour Size xQty (always draw, even if no image)
            # Use CSV colours if available, otherwise fallback to database colour
            if idx < len(component_colours):
                colour_text = component_colours[idx]
            else:
                colour_text = str(comp.get(COLUMN_NAMES['colour'], '') or '')
            size_text = str(comp.get(COLUMN_NAMES['size'], '') or '')
            qty_text = str(item_data.get(COLUMN_NAMES['quantity'], ''))
            label_text_parts = []
            if colour_text:
                label_text_parts.append(colour_text)
            if size_text:
                label_text_parts.append(size_text)
            if qty_text:
                label_text_parts.append(f"x{qty_text}")
            label_text = " ".join(label_text_parts)
            if label_text:
                self.set_xy(x, y + thumb_h + label_gap)
                self.set_font(self.base_font, 'B', 10)
                # Centered label within the thumbnail width
                self.cell(w=thumb_w, h=label_h, txt=self._t(label_text), border=0, align='C')

        # Recipient under the enlarged image box (wrapped to avoid cutting long names)
        name_y_pos = PRODUCT_IMG_Y + box_h + 3
        self.set_xy(PRODUCT_IMG_X, name_y_pos)
        # Slightly smaller font so two lines can fit if needed
        self.set_font(self.base_font, 'B', 20)
        self.multi_cell(w=PRODUCT_IMG_W, h=8, txt=self._t(item_data.get(COLUMN_NAMES['recipient'], 'N/A')), border=0, align='C')

        # Details on the right, same structure as single item, but showing pack context
        self.set_xy(DETAILS_X_START, PRODUCT_IMG_Y)
        def write_detail(label, value):
            self.set_font(self.base_font, 'B', 12)
            self.set_x(DETAILS_X_START)
            self.cell(w=40, h=8, txt=self._t(label), border=0)
            self.set_font(self.base_font, '', 12)
            self.multi_cell(w=0, h=8, txt=self._t(f": {value}"), border=0)

        # Use pack_title from Packs Database.xlsx Column AI if available, otherwise fallback to Description
        title_value = pack_title if pack_title and str(pack_title).strip() else str(main_product.get(COLUMN_NAMES['description'], 'N/A'))
        write_detail('Title', title_value)

        # Directly under Title, show Pack Name from the PACK SKU row (database column for pack name)
        if pack_name and str(pack_name).strip():
            write_detail('Pack Name', str(pack_name))
        elif isinstance(pack_product, dict):
            # Try common column names for pack name; fall back gracefully
            candidate_keys = [
                'Pack Name', 'Pack', 'Package Name', 'Package', 'PackTitle', 'Title (Pack)'
            ]
            pack_name_value = None
            for key in candidate_keys:
                if key in pack_product and str(pack_product.get(key, '')).strip():
                    pack_name_value = str(pack_product.get(key))
                    break
            # As a last resort, reuse Description from pack row (may contain name)
            if not pack_name_value and COLUMN_NAMES['description'] in pack_product:
                pack_name_value = str(pack_product.get(COLUMN_NAMES['description']))
            if pack_name_value:
                write_detail('Pack Name', pack_name_value)
        # For PACK slips, omit Colour per request; keep Size
        write_detail('Size', str(main_product.get(COLUMN_NAMES['size'], 'N/A')))
        
        # Quantity styled like "Merge Order" pill if greater than 1
        quantity_value = str(item_data.get(COLUMN_NAMES['quantity'], 'N/A'))
        try:
            qty_num = int(quantity_value) if quantity_value.isdigit() else 0
            if qty_num > 1:
                self.set_x(DETAILS_X_START)
                pill_text = self._t(f"Quantity: {quantity_value}")
                self.set_font(self.base_font, 'B', 14)
                try:
                    text_w = self.get_string_width(pill_text)
                except Exception:
                    text_w = 60
                pill_w = min(120, max(60, text_w + 12))
                self.set_fill_color(255, 0, 0)
                self.set_text_color(255, 255, 255)
                self.cell(w=pill_w, h=8, txt=pill_text, fill=True, align='C')
                self.set_text_color(0, 0, 0)
                self.ln(0)
            else:
                write_detail('Quantity', quantity_value)
        except Exception:
            write_detail('Quantity', quantity_value)

        self.ln(15)

        write_detail('Our SKU', str(item_data.get(COLUMN_NAMES['sku'], 'N/A')))  # pack sku
        write_detail('Brand', str(main_product.get(COLUMN_NAMES['brand'], 'N/A')))
        write_detail('Product Code', str(main_product.get(COLUMN_NAMES['product_code'], 'N/A')))
        write_detail('Package', f"Pack of {pack_size}")

        # Components list: show colours if provided; fallback to SKUs
        colours_raw = item_data.get(COLUMN_NAMES['component_colours'], '')
        if isinstance(colours_raw, str) and colours_raw.strip():
            comp_list_text = colours_raw
        else:
            comp_list_text = ", ".join([str(c.get(COLUMN_NAMES['db_sku'], '')) for c in component_products])
        write_detail('Components', comp_list_text)

def _safe_add_single_page(pdf: 'PDF', item_data, product_data, item_count, total_items):
    """Render a single item page even if PDF.add_packing_slip doesn't exist."""
    add_single_fn = getattr(pdf, 'add_packing_slip', None)
    if callable(add_single_fn):
        add_single_fn(item_data, product_data, item_count, total_items)
        return
    # Manual render using internal helpers
    pdf.add_page(orientation='L')
    if hasattr(pdf, '_draw_header'):
        pdf._draw_header(item_data, product_data or {}, item_count, total_items)
    if hasattr(pdf, '_draw_product_details'):
        pdf._draw_product_details(item_data, product_data or {}, total_items)

def _image_filename_from_url(url: object) -> str:
    """Extract image filename from a ProductExport URL for assets/product_images lookup."""
    if url is None or (isinstance(url, float) and pd.isna(url)):
        return ""
    s = str(url).strip()
    if not s or s.startswith("["):
        return ""
    return Path(s.replace("\\", "/")).name


_COLOUR_IMAGE_BY_UID_CACHE: dict[str, str] | None = None


def _read_product_export_csv(**kwargs) -> tuple[pd.DataFrame, str]:
    """Load ProductExport.csv trying common encodings (exports may be cp1252)."""
    last_err: Exception | None = None
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return pd.read_csv(PRODUCT_EXPORT_FILE, encoding=encoding, low_memory=False, **kwargs), encoding
        except UnicodeDecodeError as exc:
            last_err = exc
            continue
    if last_err is not None:
        raise last_err
    raise RuntimeError(f"Could not read ProductExport.csv: {PRODUCT_EXPORT_FILE}")


def _load_colour_image_basenames_by_uid() -> dict[str, str]:
    """UID -> colour image 01 basename from ProductExport (colour-specific product shot)."""
    global _COLOUR_IMAGE_BY_UID_CACHE
    if _COLOUR_IMAGE_BY_UID_CACHE is not None:
        return _COLOUR_IMAGE_BY_UID_CACHE
    out: dict[str, str] = {}
    try:
        export_df, _encoding = _read_product_export_csv(
            usecols=["UID", "colour image 01"],
            dtype={"UID": str},
        )
    except Exception:
        _COLOUR_IMAGE_BY_UID_CACHE = out
        return out
    if "UID" not in export_df.columns or "colour image 01" not in export_df.columns:
        _COLOUR_IMAGE_BY_UID_CACHE = out
        return out
    export_df = export_df[~export_df["UID"].astype(str).str.startswith("[", na=False)]
    export_df = export_df.drop_duplicates(subset=["UID"], keep="first")
    for _, row in export_df.iterrows():
        uid = str(row.get("UID", "")).strip()
        name = _image_filename_from_url(row.get("colour image 01"))
        if uid and name:
            out[uid] = name
    _COLOUR_IMAGE_BY_UID_CACHE = out
    return out


def _resolve_product_image_path(
    product_filename: object,
    sku_value: object,
    *,
    colour_img_by_uid: dict[str, str] | None = None,
) -> tuple[Path | None, str, str]:
    """Prefer colour-specific ProductExport shot, then Database product image, then SKU.ext."""
    sku_str = str(sku_value or "").strip()
    colour_map = (
        colour_img_by_uid
        if colour_img_by_uid is not None
        else _load_colour_image_basenames_by_uid()
    )
    colour01 = colour_map.get(sku_str, "") if sku_str else ""
    if colour01:
        colour_path = PRODUCT_IMAGE_FOLDER / colour01
        if colour_path.exists():
            return colour_path, "colour_image_01", colour01
    if product_filename:
        pf = PRODUCT_IMAGE_FOLDER / str(product_filename)
        if pf.exists():
            return pf, "product_image", str(product_filename)
    if sku_str:
        for ext in (".jpg", ".png", ".jpeg", ".webp"):
            candidate = PRODUCT_IMAGE_FOLDER / f"{sku_str}{ext}"
            if candidate.exists():
                return candidate, "sku_fallback", candidate.name
    return None, "none", ""


def _load_product_export_by_uid() -> dict[str, dict]:
    """Map BTC stock id (UID) to product fields when Database.xlsx has no row."""
    try:
        export_df, _enc = _read_product_export_csv(dtype={"UID": str})
    except FileNotFoundError:
        return {}
    except Exception as e:
        print(f"[PDF] Warning: Could not load ProductExport.csv fallback: {e}")
        return {}

    if "UID" not in export_df.columns:
        return {}

    export_df = export_df[~export_df["UID"].astype(str).str.startswith("[", na=False)]
    export_df = export_df.drop_duplicates(subset=["UID"], keep="first")

    by_uid: dict[str, dict] = {}
    for _, row in export_df.iterrows():
        uid = str(row.get("UID", "")).strip()
        if not uid:
            continue
        hi_res = row.get("image_url_high_res") or row.get("image_url_medium_res")
        by_uid[uid] = {
            COLUMN_NAMES["db_sku"]: uid,
            COLUMN_NAMES["product_code"]: str(row.get("SPC", "") or "").strip(),
            COLUMN_NAMES["brand"]: str(row.get("Brand", "") or "").strip(),
            COLUMN_NAMES["colour"]: str(row.get("Colour Name", "") or "").strip(),
            COLUMN_NAMES["size"]: str(row.get("Size", "") or "").strip(),
            COLUMN_NAMES["description"]: str(row.get("Description", "") or "").strip(),
            COLUMN_NAMES["package"]: "",
            COLUMN_NAMES["product_image_filename"]: _image_filename_from_url(hi_res),
            COLUMN_NAMES["brand_image_filename"]: _image_filename_from_url(row.get("brand image")),
        }
    return by_uid


def _row_to_product_dict(row) -> dict:
    if isinstance(row, pd.DataFrame):
        row = row.iloc[0]
    return row.to_dict()


def _lookup_product_details(
    products_df: pd.DataFrame,
    sku: str,
    export_by_uid: dict[str, dict] | None = None,
) -> dict:
    """Resolve product row from Database.xlsx, then ProductExport.csv (UID = stock id)."""
    if not sku or (isinstance(sku, float) and pd.isna(sku)):
        return {}
    sku_str = str(sku).strip()
    if not sku_str:
        return {}

    try:
        return _row_to_product_dict(products_df.loc[sku_str])
    except KeyError:
        pass

    fallback = (export_by_uid or {}).get(sku_str)
    if fallback:
        print(f"     Using ProductExport fallback for SKU '{sku_str}' (not in Database.xlsx)")
        return dict(fallback)
    return {}


def _load_pack_names_map(excel_path: str) -> dict:
    """Load {normalized_pack_sku -> Pack Name} using openpyxl (B=SKU, C=Pack Name)."""
    try:
        from openpyxl import load_workbook
    except Exception:
        return {}
    try:
        wb = load_workbook(excel_path, data_only=True, read_only=True)
        ws = wb.active
        pack_names = {}
        for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
            pack_sku = row[1]
            pack_name = row[2] if len(row) > 2 else None
            if not pack_sku:
                continue
            sku_str = str(pack_sku).strip()
            if not sku_str:
                continue
            normalized = sku_str.split('-')[0] if '-' in sku_str else sku_str
            pack_names[normalized] = str(pack_name).strip() if pack_name is not None else ''
        return pack_names
    except Exception:
        return {}

def _load_pack_titles_map(excel_path: str) -> dict:
    """Load {normalized_pack_sku -> Title} using openpyxl (B=SKU, AI=Title)."""
    try:
        from openpyxl import load_workbook
    except Exception:
        return {}
    try:
        wb = load_workbook(excel_path, data_only=True, read_only=True)
        ws = wb.active
        pack_titles = {}
        # Column AI = 35 (1-based), so index 34 (0-based)
        for row in ws.iter_rows(min_row=2, max_col=35, values_only=True):
            pack_sku = row[1]  # Column B
            pack_title = row[34] if len(row) > 34 else None  # Column AI
            if not pack_sku:
                continue
            sku_str = str(pack_sku).strip()
            if not sku_str:
                continue
            normalized = sku_str.split('-')[0] if '-' in sku_str else sku_str
            pack_titles[normalized] = str(pack_title).strip() if pack_title is not None else ''
        return pack_titles
    except Exception:
        return {}


def generate_packing_slips_for_tag(csv_filename, tag_id, output_path=None):
    """Generate PDF packing slips for a specific tag ID CSV file. Returns True if a PDF was saved."""
    print(f"\n[PDF] Starting PDF generation for tag {tag_id}...")

    try:
        orders_df = pd.read_csv(
            csv_filename,
            dtype={
                COLUMN_NAMES['sku']: str,
                COLUMN_NAMES['process']: str,
                COLUMN_NAMES['components']: str,
                COLUMN_NAMES['component_colours']: str,
            },
        )
        products_df = pd.read_excel(PRODUCT_DATABASE_FILE, dtype={COLUMN_NAMES['db_sku']: str, COLUMN_NAMES['product_code']: str})
        # Load Pack Names and Titles maps via openpyxl
        pack_names_map = _load_pack_names_map(PACKS_DATABASE_FILE)
        pack_titles_map = _load_pack_titles_map(PACKS_DATABASE_FILE)
        export_by_uid = _load_product_export_by_uid()
        try:
            from run_script import load_packs_database
            packs_components_map = load_packs_database(str(PACKS_DATABASE_FILE))
        except Exception:
            packs_components_map = {}
        
        products_df = products_df.drop_duplicates(subset=[COLUMN_NAMES['db_sku']], keep='first')
        products_df = products_df.set_index(COLUMN_NAMES['db_sku'])
        
        print(f"[PDF] Loaded {len(orders_df)} order items and {len(products_df)} products.")
        if export_by_uid:
            print(f"[PDF] ProductExport fallback: {len(export_by_uid)} SKUs available.")
    except FileNotFoundError as e:
        print(f"\n[ERROR] FATAL: Required file not found. Please check paths.\n    Details: {e}")
        return False
    except KeyError as e:
        print(f"\n[ERROR] FATAL: A required column is missing from a file.\n    Column Name: {e}")
        return False
    except Exception as e:
        print(f"\n[ERROR] FATAL: An unexpected error occurred during file loading: {e}")
        return False

    if orders_df.empty:
        print("[WARNING] No packing-slip rows in CSV — PDF not created.")
        return False

    saved_any = False

    # Group by tag (process) - should be all the same tag
    grouped_by_process = orders_df.groupby(COLUMN_NAMES['process'])

    for process_number, process_df in grouped_by_process:
        print(f"\n--- Processing Tag: {process_number} ---")
        
        pdf = PDF()
        pdf.set_auto_page_break(auto=True, margin=MARGIN)
        pages_before = pdf.page

        grouped_orders = process_df.groupby(COLUMN_NAMES['order_id'], sort=False)

        for order_id, items_in_order in grouped_orders:
            total_items_in_order = len(items_in_order)
            print(f"   Processing Order {order_id} ({total_items_in_order} item(s))...")

            for item_count, (_, item) in enumerate(items_in_order.iterrows(), 1):
                sku = item.get(COLUMN_NAMES['sku'])
                linked_sku = item.get('Linked SKU', '') if 'Linked SKU' in item else ''
                components_str = item.get(COLUMN_NAMES['components'], '') if COLUMN_NAMES['components'] in item else ''
                colours_str = item.get(COLUMN_NAMES['component_colours'], '') if COLUMN_NAMES['component_colours'] in item else ''
                
                # Use linked SKU for product lookup if available (for non-pack items)
                lookup_sku = linked_sku if linked_sku else sku
                
                print(f"     - Item {item_count}/{total_items_in_order}, SKU: {sku}")
                if linked_sku and linked_sku != sku:
                    print(f"       Using linked SKU for product lookup: {linked_sku}")

                if not sku or pd.isna(sku):
                    continue

                # If components present, treat as Pack of N
                components = []
                if isinstance(components_str, str) and components_str.strip():
                    components = [c.strip() for c in components_str.split(',') if c.strip()]

                # OOS / issue CSV rows omit Components — recover from Packs Database
                if not components:
                    normalized_pack_lookup = str(sku).split('-')[0] if sku and '-' in str(sku) else str(sku)
                    pack_entries = packs_components_map.get(str(normalized_pack_lookup).strip(), [])
                    if pack_entries:
                        components = [str(e.get("sku", "")).strip() for e in pack_entries if e.get("sku")]
                        if not (isinstance(colours_str, str) and colours_str.strip()):
                            colours_str = ",".join(str(e.get("colour", "") or "") for e in pack_entries)

                if components:
                    # Build product details list for each component
                    component_products = []
                    for comp_sku in components:
                        comp_details = _lookup_product_details(products_df, comp_sku, export_by_uid)
                        if not comp_details:
                            print(f"       Warning: Product details not found for component SKU '{comp_sku}'.")
                        component_products.append({**comp_details, COLUMN_NAMES['db_sku']: comp_sku})

                    # Attach colours string into item_data so renderer can show colours instead of IDs
                    item_dict = item.to_dict()
                    item_dict[COLUMN_NAMES['component_colours']] = colours_str
                    pack_product_details = _lookup_product_details(products_df, sku, export_by_uid) or None
                    normalized_pack = str(sku).split('-')[0] if sku and '-' in str(sku) else str(sku)
                    pack_name_value = pack_names_map.get(normalized_pack, '')
                    pack_title_value = pack_titles_map.get(normalized_pack, '')

                    add_pack_fn = getattr(pdf, 'add_pack_slip', None)
                    if callable(add_pack_fn):
                        add_pack_fn(
                            item_dict,
                            component_products,
                            item_count,
                            total_items_in_order,
                            pack_product=pack_product_details,
                            pack_name=pack_name_value,
                            pack_title=pack_title_value,
                        )
                    else:
                        # Fallback: render as single slip using pack SKU details
                        fallback_item = item.to_dict()
                        _safe_add_single_page(pdf, fallback_item, pack_product_details or {}, item_count, total_items_in_order)
                else:
                    product_details = _lookup_product_details(products_df, lookup_sku, export_by_uid)
                    if product_details and linked_sku and linked_sku != sku:
                        print(f"     Found product details using linked SKU: {lookup_sku}")
                    elif not product_details:
                        print(f"     Warning: Product details not found in database for SKU '{lookup_sku}' (Original: {sku}).")

                    # Even for single items (no components), enrich with Pack Name/Title from Packs DB if available
                    normalized_pack = str(sku).split('-')[0] if sku and '-' in str(sku) else str(sku)
                    pack_name_value = pack_names_map.get(normalized_pack, '')
                    pack_title_value = pack_titles_map.get(normalized_pack, '')
                    
                    # Debug: Print what we found in packs database
                    print(f"     Looking for normalized pack: {normalized_pack}")
                    print(f"     Pack name found: {pack_name_value}")
                    print(f"     Pack title found: {pack_title_value}")

                    add_single_fn = getattr(pdf, 'add_packing_slip', None)
                    if callable(add_single_fn):
                        add_single_fn(
                            item.to_dict(),
                            product_details,
                            item_count,
                            total_items_in_order,
                            pack_product=product_details or None,
                            pack_name=pack_name_value,
                            pack_title=pack_title_value,
                        )
                    else:
                        _safe_add_single_page(pdf, item.to_dict(), product_details, item_count, total_items_in_order)

        if pdf.page <= pages_before:
            print(f"   [WARNING] No PDF pages generated for tag '{process_number}' — skipping save.")
            continue

        try:
            if output_path:
                # Use provided output path
                output_filename = Path(output_path)
                output_filename.parent.mkdir(parents=True, exist_ok=True)
            else:
                pdf_output_folder = tag_output_dir("pdf_output")
                today_str = date.today().strftime("%Y-%m-%d")
                date_folder = pdf_output_folder / today_str
                date_folder.mkdir(parents=True, exist_ok=True)
                from run_script import pdf_filename_for_tag

                output_filename = date_folder / pdf_filename_for_tag(tag_id)
            
            pdf.output(str(output_filename))
            if output_filename.exists() and output_filename.stat().st_size > 0:
                saved_any = True
                print(f"   Success! PDF for tag '{tag_id}' generated at: {output_filename.resolve()}")
            else:
                print(f"   [WARNING] PDF file missing or empty after save: {output_filename}")
        except Exception as e:
            print(f"\n   ERROR: Could not save the PDF for tag '{tag_id}'. Is the file open elsewhere?\n   Reason: {e}")

    print(f"\n[PDF] PDF generation for tag {tag_id} completed!")
    return saved_any

def generate_packing_slips():
    """Main function to load data and generate PDFs based on Process Number."""
    print("\n🚀 Starting packing slip generation...")

    try:
        orders_df = pd.read_csv(PLAIN_ITEMS_CSV, dtype={COLUMN_NAMES['sku']: str, COLUMN_NAMES['process']: str})
        products_df = pd.read_excel(PRODUCT_DATABASE_FILE, dtype={COLUMN_NAMES['db_sku']: str, COLUMN_NAMES['product_code']: str})
        # Load Pack Names and Titles maps via openpyxl
        pack_names_map = _load_pack_names_map(PACKS_DATABASE_FILE)
        pack_titles_map = _load_pack_titles_map(PACKS_DATABASE_FILE)
        export_by_uid = _load_product_export_by_uid()
        
        products_df = products_df.drop_duplicates(subset=[COLUMN_NAMES['db_sku']], keep='first')
        products_df = products_df.set_index(COLUMN_NAMES['db_sku'])
        
        print(f"✅ Loaded {len(orders_df)} order items and {len(products_df)} products.")
    except FileNotFoundError as e:
        print(f"\n❌ FATAL ERROR: Required file not found. Please check paths.\n    Details: {e}")
        return
    except KeyError as e:
        print(f"\n❌ FATAL ERROR: A required column is missing from a file.\n    Column Name: {e}")
        return
    except Exception as e:
        print(f"\n❌ FATAL ERROR: An unexpected error occurred during file loading: {e}")
        return
    
    orders_df = orders_df.sort_values(by=COLUMN_NAMES['process'])
    
    grouped_by_process = orders_df.groupby(COLUMN_NAMES['process'])

    for process_number, process_df in grouped_by_process:
        print(f"\n--- 🏭 Processing Group: {process_number} ---")
        
        pdf = PDF()
        pdf.set_auto_page_break(auto=True, margin=MARGIN)

        grouped_orders = process_df.groupby(COLUMN_NAMES['order_id'], sort=False)

        for order_id, items_in_order in grouped_orders:
            total_items_in_order = len(items_in_order)
            print(f"   📦 Processing Order {order_id} ({total_items_in_order} item(s))...")

            for item_count, (_, item) in enumerate(items_in_order.iterrows(), 1):
                sku = item.get(COLUMN_NAMES['sku'])
                print(f"     - Item {item_count}/{total_items_in_order}, SKU: {sku}")

                if not sku or pd.isna(sku):
                    continue

                product_details = _lookup_product_details(products_df, sku, export_by_uid)
                if not product_details:
                    print(f"     ⚠️ Warning: Product details not found in database for SKU '{sku}'.")
                
                normalized_pack = str(sku).split('-')[0] if sku and '-' in str(sku) else str(sku)
                pack_name_value = pack_names_map.get(normalized_pack, '')
                pack_title_value = pack_titles_map.get(normalized_pack, '')
                pdf.add_packing_slip(item.to_dict(), product_details, item_count, total_items_in_order, pack_product=product_details, pack_name=pack_name_value, pack_title=pack_title_value)

        try:
            pdf_output_folder = tag_output_dir("pdf_output")
            today_str = date.today().strftime("%Y-%m-%d")
            date_folder = pdf_output_folder / today_str
            date_folder.mkdir(parents=True, exist_ok=True)
            
            actual_process_id = str(process_number).split()[-1]
            output_filename = date_folder / f"{actual_process_id}.pdf"
            
            pdf.output(str(output_filename))
            print(f"   🎉 Success! PDF for process '{process_number}' generated at: {output_filename.resolve()}")
        except Exception as e:
            print(f"\n   ❌ FATAL ERROR: Could not save the PDF for process '{process_number}'. Is the file open elsewhere?\n   Reason: {e}")

    print("\n\n✅ All PDF files have been generated successfully!")

if __name__ == "__main__":
    print("PDF Generator - Packing Slips")
    print("=" * 50)
    print("Make sure you have:")
    print(f"1. {PLAIN_ITEMS_CSV} - Orders data")
    print(f"2. {PRODUCT_DATABASE_FILE} - Product database")
    print(f"3. {PRODUCT_IMAGE_FOLDER} - Product images folder")
    print(f"4. {BRAND_IMAGE_FOLDER} - Brand logos folder")
    print()
    
    generate_packing_slips()
