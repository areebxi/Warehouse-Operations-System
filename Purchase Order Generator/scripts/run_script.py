"""
Simple script to run the ShipStation orders fetcher
This script uses the config.py file for API credentials
"""

import app_paths  # noqa: F401 — configures import paths before other local imports

from shipstation_orders import ShipStationAPI
from config import SHIPSTATION_API_KEY, SHIPSTATION_API_SECRET
from pdf_generator import generate_packing_slips_for_tag
import sys
import os
import csv
import json
import socket
from datetime import datetime
from ftplib import FTP, error_perm, error_temp, error_reply

from app_paths import data_path, packs_database_path, shipstation_tags_path, tag_output_dir
from stock_resolver import (
    NOT_FOUND_STATUSES,
    STATUS_NOT_FOUND,
    load_custom_label_stock_map,
    not_found_status,
    resolve_stock_level,
)


def get_process_no_for_tag(tag_id: str):
    """Read Process No from ShipStation Tags.xlsx (C: Tag ID, D: Process No)."""
    try:
        try:
            import openpyxl
        except Exception:
            print("[WARNING] openpyxl not available to read ShipStation Tags.xlsx")
            return None

        xlsx_path = str(shipstation_tags_path())
        if not os.path.exists(xlsx_path):
            print(f"[WARNING] ShipStation Tags.xlsx not found at: {xlsx_path}")
            return None

        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = wb.active

        search_value = str(tag_id).strip()
        for row in ws.iter_rows(min_row=1):
            tag_cell = row[2]  # Column C (0-based index 2)
            process_cell = row[3]  # Column D (0-based index 3)
            tag_val = "" if tag_cell.value is None else str(tag_cell.value).strip()
            if tag_val == search_value:
                process_val = None if process_cell.value is None else str(process_cell.value).strip()
                return process_val or None
        return None
    except Exception as e:
        print(f"[WARNING] Failed to read Process No from ShipStation Tags.xlsx: {e}")
        return None


def pdf_filename_for_tag(tag_id: str, process_no: str | None = None) -> str:
    if process_no is None:
        process_no = get_process_no_for_tag(tag_id)
    if process_no:
        return f"{process_no}.pdf"
    return f"Tag_{tag_id}.pdf"


# --- FTP Configuration ---
def load_packs_database(excel_path=None):
    """
    Load packs mapping from Excel.
    - Column B: pack SKU to search by
    - Columns E, H, K, N, Q: component SKUs to try

    Returns dict: { pack_sku -> [component_sku, ...] }
    """
    packs_map = {}
    try:
        try:
            from openpyxl import load_workbook
        except ImportError:
            print("[WARNING] openpyxl not installed. Skipping Packs Database lookup.")
            return packs_map

        if excel_path is None:
            excel_path = str(packs_database_path())
        if not os.path.exists(excel_path):
            print(f"[WARNING] Packs Database not found at: {excel_path}")
            return packs_map

        wb = load_workbook(excel_path, data_only=True, read_only=True)
        ws = wb.active

        # Column indices (1-based):
        #   Pack SKU in B=2
        #   Component SKUs in E=5, H=8, K=11, N=14, Q=17
        #   Their Colours in     F=6, I=9, L=12, O=15, R=18 (next column)
        pack_col_idx = 2
        component_col_indices = [5, 8, 11, 14, 17]
        colour_col_indices =   [6, 9, 12, 15, 18]

        # Iterate efficiently with values_only
        for row in ws.iter_rows(min_row=2, max_col=18, values_only=True):
            pack_sku = row[pack_col_idx - 1]
            if not pack_sku:
                continue
            pack_sku = str(pack_sku).strip()
            if not pack_sku:
                continue

            component_entries = []
            for comp_col_idx, colour_col_idx in zip(component_col_indices, colour_col_indices):
                val = row[comp_col_idx - 1] if comp_col_idx - 1 < len(row) else None
                if val is None:
                    continue
                sku_str = str(val).strip()
                if not sku_str:
                    continue
                if '-' in sku_str:
                    sku_str = sku_str.split('-')[0]
                colour_val = row[colour_col_idx - 1] if colour_col_idx - 1 < len(row) else None
                colour_str = str(colour_val).strip() if colour_val is not None else ''
                component_entries.append({"sku": sku_str, "colour": colour_str})

            if component_entries:
                normalized_pack = pack_sku.split('-')[0] if '-' in pack_sku else pack_sku
                packs_map[normalized_pack] = component_entries

        print(f"[INFO] Packs Database loaded: {len(packs_map)} pack SKUs mapped")
        return packs_map
    except Exception as e:
        print(f"[WARNING] Failed to load Packs Database: {e}")
        return {}

def load_pack_names(excel_path=None):
    """
    Load Pack Name mapping from Excel.
    - Column B: pack SKU to search by
    - Column C: Pack Name

    Returns dict: { normalized_pack_sku -> pack_name }
    """
    pack_names = {}
    try:
        try:
            from openpyxl import load_workbook
        except ImportError:
            print("[WARNING] openpyxl not installed. Skipping Pack Name lookup.")
            return pack_names

        if excel_path is None:
            excel_path = str(packs_database_path())
        if not os.path.exists(excel_path):
            print(f"[WARNING] Packs Database not found at: {excel_path}")
            return pack_names

        wb = load_workbook(excel_path, data_only=True, read_only=True)
        ws = wb.active

        # Column indices (1-based): Pack SKU in B=2, Pack Name in C=3
        for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
            pack_sku = row[1]
            pack_name = row[2] if len(row) > 2 else None
            if not pack_sku:
                continue
            sku_str = str(pack_sku).strip()
            if not sku_str:
                continue
            normalized_pack = sku_str.split('-')[0] if '-' in sku_str else sku_str
            pack_names[normalized_pack] = str(pack_name).strip() if pack_name is not None else ''

        print(f"[INFO] Pack Names loaded: {len(pack_names)} entries")
        return pack_names
    except Exception as e:
        print(f"[WARNING] Failed to load Pack Names: {e}")
        return {}


def _pack_key(sku: str) -> str:
    sku = (sku or "").strip()
    return sku.split("-", 1)[0] if "-" in sku else sku


def is_discount_line_item(item) -> bool:
    """True for Etsy/marketplace discount adjustment lines (no stock check)."""
    return (item.get("name") or "").strip().casefold() == "discount"


def basic_sku(original_sku: str) -> str:
    """Marketplace prefix: part before the first dash."""
    return _pack_key(original_sku)


def build_issue_row(
    order_number,
    recipient_name,
    quantity,
    original_sku,
    tag_id,
    stock_level,
    process_no,
    *,
    item_sku="",
    stock_id="",
    status="",
) -> list:
    """Internal row for stock-issue exports (9 fields, or 10 when status set).

    Item SKU is blank when the before-dash prefix was not a real stock hit
    (custom-label / after-dash path, or not found). Complete SKU is always
    the full marketplace / ShipStation SKU. Status is set for not-found rows so
    the CSV can distinguish custom-label vs stock-levels misses.
    """
    row = [
        order_number,
        recipient_name,
        quantity,
        item_sku or "",
        (original_sku or "").strip(),
        stock_id or "",
        tag_id,
        stock_level,
        process_no,
    ]
    if status:
        row.append(status)
    return row


def _issue_item_sku(effective: str, used_fallback: bool) -> str:
    """Display Item SKU only for a real primary stock-id hit."""
    if used_fallback:
        return ""
    return effective or ""


def load_stock_levels(log=print) -> dict[str, int]:
    """Load stock levels CSV into a stock_id → quantity map."""
    remote_file, stock_path, stock_file_name = _stock_file_paths()
    stock_levels: dict[str, int] = {}
    log(
        f"[STOCK] Loading stock levels from {stock_file_name} "
        f"(local: {stock_path}, remote: {remote_file})"
    )
    try:
        with open(stock_path, "r", encoding="utf-8") as stock_file:
            stock_reader = csv.DictReader(stock_file)
            for row in stock_reader:
                stock_id = row.get("stock_id", "").strip()
                free_stock = int(row.get("free_stock", 0))
                stock_levels[stock_id] = free_stock
        log(f"[SUCCESS] Loaded stock levels for {len(stock_levels)} items from {stock_file_name}")
    except FileNotFoundError:
        log(
            f"[WARNING] {stock_file_name} not found at {stock_path}. "
            "Proceeding without stock checks. "
            "Set FTP_LOCAL_FILE in config.py or place the file in data/."
        )
    except Exception as e:
        log(
            f"[WARNING] Error reading {stock_file_name} at {stock_path}: {e}. "
            "Proceeding without stock checks."
        )
    return stock_levels


def validate_orders_stock(
    filtered_orders,
    tag_id: str,
    process_no,
    stock_levels: dict[str, int],
    packs_map: dict,
    pack_names_map: dict,
    custom_label_map: dict[str, str],
    log=print,
    labels_missing_stock_id: set[str] | None = None,
):
    """
    Build in-stock, out-of-stock, and not-found item lists (pack-aware, atomic per order).
    In-stock single-SKU rows: 8 fields ending with marketplace_sku.
    In-stock pack rows: 11 fields ending with marketplace_sku.
    """
    missing_label_ids = labels_missing_stock_id or set()
    in_stock_items = []
    out_of_stock_items = []
    not_found_items = []

    for order in filtered_orders:
        order_number = order.get("orderNumber", "")
        ship_to = order.get("shipTo") or {}
        recipient_name = ship_to.get("name", "")

        raw_items = order.get("items", []) or order.get("lineItems", [])
        items = []
        for item in raw_items:
            if is_discount_line_item(item):
                log(f"[SKIP] Ignoring Discount line on order {order_number}")
            else:
                items.append(item)

        if items:
            order_ok = True
            in_rows_for_order = []
            out_rows_for_order = []
            missing_rows_for_order = []
            for item in items:
                quantity = item.get("quantity", 1)
                original_sku = item.get("sku", "")
                pack_key = _pack_key(original_sku)

                component_candidates = packs_map.get(pack_key, [])
                if component_candidates:
                    component_levels = {}
                    effective_components = []
                    any_missing = False
                    insufficient = False
                    pack_nf_status = STATUS_NOT_FOUND
                    pack_nf_stock_id = ""
                    for comp_entry in component_candidates:
                        comp_original = comp_entry.get("sku") or ""
                        level, effective, _marketplace, used_fallback = resolve_stock_level(
                            comp_original, stock_levels, custom_label_map
                        )
                        component_levels[effective] = level
                        effective_components.append(effective)
                        if used_fallback:
                            log(
                                f"[CUSTOM LABEL] Component {comp_original} -> stock ID {effective}"
                            )
                        if level == -1:
                            if not any_missing:
                                pack_nf_status = not_found_status(
                                    comp_original,
                                    custom_label_map,
                                    missing_label_ids,
                                    used_fallback=used_fallback,
                                )
                                pack_nf_stock_id = effective if used_fallback else ""
                            any_missing = True
                        elif level < quantity:
                            insufficient = True

                    if any_missing:
                        missing_rows_for_order.append(
                            build_issue_row(
                                order_number,
                                recipient_name,
                                quantity,
                                original_sku,
                                tag_id,
                                -1,
                                process_no,
                                item_sku=pack_key,
                                stock_id=pack_nf_stock_id,
                                status=pack_nf_status,
                            )
                        )
                        log(
                            f"[ERROR] Pack components missing in stock file: Order {order_number}, "
                            f"Pack {pack_key} -> {component_candidates} ({pack_nf_status})"
                        )
                        order_ok = False
                    elif insufficient:
                        log(
                            f"[WARNING] Pack insufficient stock: Order {order_number}, Pack {pack_key} "
                            f"(Need {quantity} each) -> {component_levels}"
                        )
                        order_ok = False
                        out_rows_for_order.append(
                            build_issue_row(
                                order_number,
                                recipient_name,
                                quantity,
                                original_sku,
                                tag_id,
                                0,
                                process_no,
                                item_sku=pack_key,
                            )
                        )
                    else:
                        components_joined = ",".join(effective_components)
                        colours_joined = ",".join(
                            [c.get("colour", "") for c in component_candidates]
                        )
                        pack_name_value = pack_names_map.get(pack_key, "")
                        in_rows_for_order.append(
                            [
                                order_number,
                                recipient_name,
                                quantity,
                                pack_key,
                                pack_name_value,
                                tag_id,
                                1,
                                components_joined,
                                colours_joined,
                                process_no,
                                "",
                            ]
                        )
                else:
                    stock_level, effective, marketplace, used_fallback = resolve_stock_level(
                        original_sku, stock_levels, custom_label_map
                    )
                    if used_fallback:
                        log(
                            f"[CUSTOM LABEL] {original_sku} -> stock ID {effective} "
                            f"(label {original_sku.split('-', 1)[1].strip()})"
                        )
                    # Packing/EDI use resolved stock ID; marketplace SKU holds full original.
                    packing_sku = effective if effective else ""
                    item_data = [
                        order_number,
                        recipient_name,
                        quantity,
                        packing_sku,
                        tag_id,
                        stock_level,
                        process_no,
                        marketplace or (original_sku or "").strip(),
                    ]
                    if stock_level == -1:
                        status = not_found_status(
                            original_sku,
                            custom_label_map,
                            missing_label_ids,
                            used_fallback=used_fallback,
                        )
                        missing_rows_for_order.append(
                            build_issue_row(
                                order_number,
                                recipient_name,
                                quantity,
                                original_sku,
                                tag_id,
                                -1,
                                process_no,
                                item_sku="",
                                stock_id=effective if used_fallback else "",
                                status=status,
                            )
                        )
                        log(
                            f"[ERROR] {status}: Order {order_number}, "
                            f"SKU {(original_sku or '').strip() or effective}"
                        )
                        order_ok = False
                    elif stock_level < quantity:
                        log(
                            f"[WARNING] Insufficient stock: Order {order_number}, SKU {effective} "
                            f"(Have {stock_level}, Need {quantity})"
                        )
                        order_ok = False
                        out_rows_for_order.append(
                            build_issue_row(
                                order_number,
                                recipient_name,
                                quantity,
                                original_sku,
                                tag_id,
                                stock_level,
                                process_no,
                                item_sku=_issue_item_sku(effective, used_fallback),
                                stock_id=effective,
                            )
                        )
                    else:
                        in_rows_for_order.append(item_data)

            if order_ok:
                in_stock_items.extend(in_rows_for_order)
                log(
                    f"[FOUND] Order {order_number} fully in stock ({len(in_rows_for_order)} lines)"
                )
            else:
                out_of_stock_items.extend(out_rows_for_order)
                if missing_rows_for_order:
                    not_found_items.extend(missing_rows_for_order)
                log(f"[WARNING] Order {order_number} moved to out-of-stock (atomic rule)")
        else:
            log(f"[WARNING] Order {order_number} has no items")
            in_stock_items.append(
                [order_number, recipient_name, "", "", tag_id, "N/A", process_no, ""]
            )

    return in_stock_items, out_of_stock_items, not_found_items


def normalize_packing_rows(in_stock_items):
    """Normalize in-stock rows to packing list CSV format (11 columns)."""
    normalized_rows = []
    for row in in_stock_items:
        # Not-found issue rows may include Status as a 10th field.
        if len(row) == 10 and str(row[9]) in NOT_FOUND_STATUSES:
            row = row[:9]
        if len(row) == 8:
            order_number, recipient_name, quantity, sku, tag_id, stock_level, pno, marketplace = row
            normalized_rows.append(
                [
                    order_number,
                    recipient_name,
                    quantity,
                    sku,
                    "",
                    tag_id,
                    stock_level,
                    "",
                    "",
                    pno,
                    marketplace,
                ]
            )
        elif len(row) == 11:
            normalized_rows.append(row)
        elif len(row) == 10:
            normalized_rows.append(list(row) + [""])
        elif len(row) == 9:
            (
                order_number,
                recipient_name,
                quantity,
                item_sku,
                complete_sku,
                stock_id,
                tag_id,
                stock_level,
                pno,
            ) = row
            packing_sku = stock_id if stock_id else item_sku
            normalized_rows.append(
                [
                    order_number,
                    recipient_name,
                    quantity,
                    packing_sku,
                    "",
                    tag_id,
                    stock_level,
                    "",
                    "",
                    pno,
                    complete_sku,
                ]
            )
        elif len(row) == 7:
            order_number, recipient_name, quantity, sku, tag_id, stock_level, pno = row
            normalized_rows.append(
                [
                    order_number,
                    recipient_name,
                    quantity,
                    sku,
                    "",
                    tag_id,
                    stock_level,
                    "",
                    "",
                    pno,
                    "",
                ]
            )
        else:
            normalized_rows.append(row)
    return normalized_rows


def rows_for_pdf_slips(
    in_stock_items,
    out_of_stock_items=None,
    not_found_items=None,
    packs_map=None,
    pack_names_map=None,
):
    """Packing-slip PDF rows for EDI-eligible (in-stock) orders.

    Callers should pass only in_stock_items (same set as the EDI file).
    Optional OOS / not-found args are kept for compatibility but are unused when
    empty. Pack Components are rehydrated from Packs Database when missing.
    """
    issue_rows = list(out_of_stock_items or []) + list(not_found_items or [])
    rows = normalize_packing_rows((in_stock_items or []) + issue_rows)
    if not packs_map:
        return rows

    enriched = []
    for row in rows:
        if len(row) < 11:
            enriched.append(row)
            continue
        row = list(row)
        components_val = str(row[7] or "").strip()
        if not components_val:
            pack_key = _pack_key(str(row[3] or ""))
            comps = packs_map.get(pack_key) or []
            if comps:
                row[7] = ",".join(str(c.get("sku", "") or "") for c in comps)
                row[8] = ",".join(str(c.get("colour", "") or "") for c in comps)
                if not str(row[4] or "").strip() and pack_names_map:
                    row[4] = pack_names_map.get(pack_key, "") or ""
        enriched.append(row)
    return enriched


def write_packing_list_csv(filename: str, in_stock_items) -> None:
    with open(filename, "w", newline="", encoding="utf-8") as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(
            [
                "Order",
                "Recipient",
                "Quantity",
                "Item SKU",
                "Pack Name",
                "Tag",
                "Stock Level",
                "Components",
                "Component Colours",
                "Process No",
                "Marketplace SKU",
            ]
        )
        writer.writerows(normalize_packing_rows(in_stock_items))


STATUS_OUT_OF_STOCK = "Out of Stock"


def write_stock_issues_csv(filename: str, out_of_stock_items, not_found_items) -> None:
    """Write one combined CSV with specific Not Found reasons or Out of Stock."""
    with open(filename, "w", newline="", encoding="utf-8") as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(
            [
                "Order",
                "Recipient",
                "Quantity",
                "Item SKU",
                "Complete SKU",
                "Stock ID",
                "Tag",
                "Stock Level",
                "Process No",
                "Status",
            ]
        )
        for row in not_found_items or []:
            status = STATUS_NOT_FOUND
            if len(row) >= 10 and row[9]:
                status = row[9]
                row = row[:9]
            (
                order_number,
                recipient_name,
                quantity,
                item_sku,
                complete_sku,
                stock_id,
                tag_id,
                _stock_level,
                process_no,
            ) = row
            writer.writerow(
                [
                    order_number,
                    recipient_name,
                    quantity,
                    item_sku,
                    complete_sku,
                    stock_id,
                    tag_id,
                    "N/A",
                    process_no,
                    status,
                ]
            )
        for row in out_of_stock_items or []:
            if len(row) >= 10:
                row = row[:9]
            (
                order_number,
                recipient_name,
                quantity,
                item_sku,
                complete_sku,
                stock_id,
                tag_id,
                stock_level,
                process_no,
            ) = row
            writer.writerow(
                [
                    order_number,
                    recipient_name,
                    quantity,
                    item_sku,
                    complete_sku,
                    stock_id,
                    tag_id,
                    stock_level,
                    process_no,
                    STATUS_OUT_OF_STOCK,
                ]
            )


def _unique_complete_skus(issue_rows) -> list[str]:
    seen = set()
    ordered: list[str] = []
    for row in issue_rows or []:
        if len(row) < 5:
            continue
        sku = str(row[4] or "").strip()
        if not sku or sku in seen:
            continue
        seen.add(sku)
        ordered.append(sku)
    return ordered


def format_run_summary(
    tag_label: str,
    orders_processed: int,
    in_stock_items,
    out_of_stock_items,
    not_found_items,
    issues_filename: str | None = None,
) -> str:
    """Build the end-of-run summary text for GUI log / CLI."""
    not_found_skus = _unique_complete_skus(not_found_items)
    out_of_stock_skus = _unique_complete_skus(out_of_stock_items)
    in_count = len(in_stock_items or [])
    lines = [
        "========== RUN SUMMARY ==========",
        f"Tag: {tag_label}",
        f"Orders processed: {orders_processed}",
    ]
    if not not_found_skus and not out_of_stock_skus:
        lines.append("All orders found and in stock.")
    else:
        lines.append(f"In stock: {in_count} order line(s)")
        lines.append(f"Not found: {len(not_found_skus)} SKU(s)")
        for sku in not_found_skus:
            lines.append(f"  - {sku}")
        lines.append(f"Out of stock: {len(out_of_stock_skus)} SKU(s)")
        for sku in out_of_stock_skus:
            lines.append(f"  - {sku}")
        if issues_filename:
            lines.append(f"Stock issues file: {os.path.basename(issues_filename)}")
    lines.append("=================================")
    return "\n".join(lines)


def write_edi_orders_csv(filename: str, in_stock_items, process_no) -> None:
    current_dt = datetime.now()
    date_part = current_dt.strftime("%d-%m-%Y")
    order_id_suffix = f"{date_part}-EDI-DaataaDirect"
    order_id_value = f"{process_no}-{order_id_suffix}" if process_no else order_id_suffix

    # Plain UTF-8 (no BOM): BTC's importer treats BOM as part of the first header
    # name ("\ufeffstock-id"), so stock-id is not recognized.
    with open(filename, "w", newline="", encoding="utf-8") as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(
            [
                "stock-id",
                "order-id",
                "quantity-purchased",
                "product-name",
                "recipient-name",
                "sku",
                "ship-address-1",
                "ship-address-2",
                "ship-address-3",
                "ship-city",
                "ship-state",
                "ship-postal-code",
                "ship-country",
                "collection",
                "plain-cover",
                "delivery-tracking-email",
                "delivery-tracking-sms",
                "line-note",
            ]
        )

        for item in in_stock_items:
            if len(item) == 8:
                _order_number, _recipient, quantity, sku, _tag, _level, pno, _marketplace = item
                component_list = []
            elif len(item) >= 10:
                (
                    _order_number,
                    _recipient,
                    quantity,
                    sku,
                    _pack_name,
                    _tag,
                    _level,
                    components_joined,
                    _colours,
                    pno,
                    *_rest,
                ) = item
                component_list = [c for c in (components_joined or "").split(",") if c]
            else:
                _order_number, _recipient, quantity, sku, _tag, _level, pno = item[:7]
                component_list = []

            target_skus = component_list if component_list else [sku]
            for target_sku in target_skus:
                if not target_sku:
                    continue
                writer.writerow(
                    [
                        target_sku,
                        order_id_value,
                        quantity,
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "GB",
                        "1",
                        "",
                        "",
                        "",
                        "",
                    ]
                )


def _ftp_settings():
    """Resolve FTP settings from config.py with safe defaults."""
    defaults = {
        "FTP_HOST": "ftpdata.btcactivewear.co.uk",
        "FTP_USER": "daa0001",
        "FTP_PASS": "#T/Yn7pePnPC",
        "FTP_PORT": 21,
        "FTP_PROTOCOL": "ftp",
        "FTP_REMOTE_FILE": "WebData/stock_levels_stock_id_fully_quoted.csv",
        "FTP_LOCAL_FILE": "stock_levels_stock_id_fully_quoted.csv",
        "FTP_TIMEOUT_SECONDS": 30,
        "FTP_PROBE_TIMEOUT_SECONDS": 5,
        "FTP_PASSIVE_MODE": True,
        "FTP_MAX_RETRIES": 3,
    }
    try:
        import config as app_config

        for key, default in defaults.items():
            defaults[key] = getattr(app_config, key, default)
    except Exception:
        pass
    return defaults


def _stock_file_paths(settings=None):
    """Remote path on BTC server, local path under data/, and local filename."""
    settings = settings or _ftp_settings()
    remote_file = settings["FTP_REMOTE_FILE"]
    local_basename = os.path.basename(settings["FTP_LOCAL_FILE"])
    local_path = settings["FTP_LOCAL_FILE"]
    if not os.path.isabs(local_path):
        local_path = str(data_path(local_basename))
    return remote_file, local_path, local_basename


def _ftp_probe_tcp(host: str, port: int, timeout_seconds: int) -> bool:
    """Quick TCP check before full FTP attempts."""
    try:
        with socket.create_connection((host, port), timeout=timeout_seconds):
            return True
    except OSError:
        return False


def _stock_transfer_protocol(settings: dict, ftp_port: int) -> str:
    protocol = str(settings.get("FTP_PROTOCOL", "ftp")).strip().lower()
    if protocol in ("ftp", "sftp"):
        return protocol
    if ftp_port in (22, 2022):
        return "sftp"
    return "ftp"


def _stock_transfer_blocked_message(host: str, port: int, protocol: str) -> str:
    _, local_path, local_name = _stock_file_paths()
    manual_hint = (
        f"or manually download {local_name} into the data/ folder "
        f"(set FTP_LOCAL_FILE in config.py)."
    )
    if protocol == "sftp":
        return (
            f"[ERROR] Cannot reach {host}:{port} (timed out). "
            f"Your network or firewall is blocking outbound SFTP (TCP port {port}). "
            f"Ask IT to allow port {port} to {host}, use a BTC-approved VPN, "
            f"{manual_hint}"
        )
    return (
        f"[ERROR] Cannot reach {host}:{port} (timed out). "
        f"Your network or firewall is blocking outbound FTP (TCP port {port}). "
        f"Ask IT to allow port {port} to {host}, use a BTC-approved VPN, "
        f"{manual_hint}"
    )


def _download_sftp_file(
    host: str,
    port: int,
    username: str,
    password: str,
    remote_file: str,
    local_filename: str,
    timeout_seconds: int,
    max_retries: int,
    log=print,
) -> bool:
    try:
        import paramiko
    except ImportError:
        log(
            "[ERROR] SFTP download requires the 'paramiko' package. "
            "Run: pip install paramiko"
        )
        return False

    last_error = None
    for attempt in range(1, max_retries + 1):
        transport = None
        sftp = None
        try:
            if attempt > 1:
                log(f"[INFO] SFTP retry {attempt}/{max_retries}...")
            transport = paramiko.Transport((host, port))
            transport.banner_timeout = timeout_seconds
            transport.connect(username=username, password=password)
            sftp = paramiko.SFTPClient.from_transport(transport)
            log(f"[SUCCESS] Connected to SFTP server '{host}'.")
            sftp.get(remote_file, local_filename)
            log(f"[SUCCESS] Downloaded {remote_file} → {local_filename}")
            return True
        except Exception as e:
            last_error = e
            log(f"[ERROR] SFTP attempt {attempt}/{max_retries} failed: {e}")
        finally:
            if sftp is not None:
                try:
                    sftp.close()
                except Exception:
                    pass
            if transport is not None:
                try:
                    transport.close()
                except Exception:
                    pass

    if last_error:
        log(f"[ERROR] Last SFTP error: {last_error}")
    return False


def _log_cached_stock_file(local_filename: str, log=print) -> bool:
    if not os.path.isfile(local_filename):
        return False
    modified = datetime.fromtimestamp(os.path.getmtime(local_filename))
    age_hours = (datetime.now() - modified).total_seconds() / 3600
    size_kb = os.path.getsize(local_filename) // 1024
    log(
        f"[WARNING] Using cached {local_filename} from {modified:%Y-%m-%d %H:%M} "
        f"({age_hours:.1f} hours old, {size_kb} KB). Stock checks may be outdated."
    )
    return True


def download_ftp_file(log=print):
    settings = _ftp_settings()
    ftp_host = settings["FTP_HOST"]
    ftp_user = settings["FTP_USER"]
    ftp_pass = settings["FTP_PASS"]
    ftp_port = int(settings["FTP_PORT"])
    server_filename, local_filename, local_basename = _stock_file_paths(settings)
    timeout_seconds = int(settings["FTP_TIMEOUT_SECONDS"])
    probe_timeout_seconds = int(settings.get("FTP_PROBE_TIMEOUT_SECONDS", 5))
    passive_mode = bool(settings["FTP_PASSIVE_MODE"])
    max_retries = max(1, int(settings["FTP_MAX_RETRIES"]))
    protocol = _stock_transfer_protocol(settings, ftp_port)

    if protocol == "ftp" and ftp_port == 22:
        log(
            "[WARNING] Port 22 is SFTP, not FTP. "
            "Set FTP_PROTOCOL = 'sftp' and FTP_HOST = 'sftpgo.btcactivewear.co.uk', "
            "or use FTP_PORT = 21 with ftpdata.btcactivewear.co.uk."
        )
        protocol = "sftp"

    log_label = "SFTP" if protocol == "sftp" else "FTP"
    log(f"[INFO] Downloading stock file via {log_label}...")
    log(
        f"[INFO] Stock file: remote={server_filename} → local={local_filename} "
        f"(config: FTP_REMOTE_FILE / FTP_LOCAL_FILE in config.py)"
    )
    log(f"[INFO] {log_label} host: {ftp_host}:{ftp_port} (timeout={timeout_seconds}s)")

    log(f"[INFO] Checking TCP reachability to {ftp_host}:{ftp_port}...")
    if not _ftp_probe_tcp(ftp_host, ftp_port, probe_timeout_seconds):
        log(_stock_transfer_blocked_message(ftp_host, ftp_port, protocol))
        has_cache = _log_cached_stock_file(local_filename, log=log)
        if has_cache:
            log(f"[INFO] Continuing with cached {os.path.basename(local_filename)} for stock checks.")
        else:
            log(f"[WARNING] No cached {os.path.basename(local_filename)} found. Stock checks will be skipped.")
        return False

    if protocol == "sftp":
        if _download_sftp_file(
            ftp_host,
            ftp_port,
            ftp_user,
            ftp_pass,
            server_filename,
            local_filename,
            timeout_seconds,
            max_retries,
            log=log,
        ):
            return True
        _log_cached_stock_file(local_filename, log=log)
        return False

    log(f"[INFO] FTP passive mode: {passive_mode}")
    last_error = None
    for attempt in range(1, max_retries + 1):
        ftp = None
        try:
            if attempt > 1:
                log(f"[INFO] FTP retry {attempt}/{max_retries}...")
            ftp = FTP(timeout=timeout_seconds)
            ftp.connect(ftp_host, ftp_port, timeout=timeout_seconds)
            ftp.login(ftp_user, ftp_pass)
            ftp.set_pasv(passive_mode)
            log(f"[SUCCESS] Connected to FTP server '{ftp_host}'.")

            with open(local_filename, "wb") as f:
                ftp.retrbinary(f"RETR {server_filename}", f.write)
            log(f"[SUCCESS] Downloaded {server_filename} → {local_filename}")
            return True
        except (error_perm, error_temp, error_reply, socket.timeout, TimeoutError, ConnectionRefusedError, OSError) as e:
            last_error = e
            log(f"[ERROR] FTP attempt {attempt}/{max_retries} failed: {e}")
        except FileNotFoundError:
            log(f"[ERROR] Cannot write local file '{local_filename}'. Check folder permissions.")
            return False
        except Exception as e:
            last_error = e
            log(f"[ERROR] Unexpected FTP error (attempt {attempt}/{max_retries}): {e}")
        finally:
            if ftp is not None:
                try:
                    ftp.quit()
                except Exception as e:
                    log(f"[WARNING] Error in FTP disconnection: {e}")

    log(
        f"[ERROR] Could not download free stock via FTP. "
        f"Port {ftp_port} to {ftp_host} must be allowed on your network/firewall "
        "(or use VPN if required by BTC)."
    )
    if last_error:
        log(f"[ERROR] Last FTP error: {last_error}")
    _log_cached_stock_file(local_filename, log=log)
    return False

def main():
    """
    Main function using config file for credentials
    """
    print("ShipStation Awaiting Dispatch Orders Fetcher")
    print("Tag Filtering for Awaiting Dispatch Orders Only")
    print("=" * 50)
    print(f"Started at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print()
    
    # Step 1: Download stock levels CSV from FTP/SFTP
    remote_file, local_path, _ = _stock_file_paths()
    print("[STEP 1] Downloading stock levels from BTC...")
    print(f"[INFO] Stock file: remote={remote_file} → local={local_path}")
    if not download_ftp_file():
        print("[WARNING] FTP download failed. Continuing with cached/missing stock data if available.")
    print()
    
    # Check if credentials are configured
    if SHIPSTATION_API_KEY == "your_api_key_here" or SHIPSTATION_API_SECRET == "your_api_secret_here":
        print("[ERROR] Please configure your API credentials in config.py")
        print("   Copy config_example.py to config.py and update with your actual credentials")
        sys.exit(1)
    
    # Initialize API client
    try:
        print("[INFO] Initializing API client...")
        shipstation = ShipStationAPI(SHIPSTATION_API_KEY, SHIPSTATION_API_SECRET)
        print("[SUCCESS] API client initialized successfully")
    except Exception as e:
        print(f"[ERROR] Error initializing API client: {e}")
        print("   Please check your API credentials and internet connection")
        sys.exit(1)
    
    # Set to awaiting dispatch orders only
    status_display = "Awaiting Dispatch"
    print(f"[SUCCESS] Processing: {status_display} orders only")
    
    # Get tag ID from user
    print("\n[TAG SEARCH] Tag ID Search")
    tag_id = input("Enter TAG ID: ").strip()
    if not tag_id:
        print("[ERROR] Tag ID cannot be empty!")
        sys.exit(1)
    
    print(f"\n[FETCH] Fetching {status_display.lower()} orders with tag ID: {tag_id}...")
    try:
        orders = shipstation.get_awaiting_dispatch_orders()
    except Exception as e:
        print(f"[ERROR] Error fetching orders: {e}")
        print("   This could be due to:")
        print("   - Invalid API credentials")
        print("   - Network connectivity issues")
        print("   - ShipStation API rate limiting")
        sys.exit(1)
    
    if not orders:
        print(f"[INFO] No {status_display.lower()} orders found.")
        print("   This means all your orders are in other statuses.")
        return
    
    # Filter orders by tag ID
    print("[FILTER] Filtering orders by tag ID...")
    original_count = len(orders)
    filtered_orders = []
    
    for order in orders:
        order_tags = order.get('tagIds') or []  # Handle None values
        
        # Check if order has the specified tag ID
        tag_id_str = str(tag_id)
        order_tags_str = [str(tag) for tag in order_tags]
        
        if tag_id_str in order_tags_str:
            filtered_orders.append(order)
            print(f"[FOUND] Found order {order.get('orderNumber', 'N/A')} with tag ID {tag_id}")
    
    print(f"[FILTER] Filtered {original_count} orders down to {len(filtered_orders)} orders")
    
    if not filtered_orders:
        print("[INFO] No orders found with the specified tag ID.")
        print("   Available tag IDs in your orders:")
        all_tag_ids = set()
        for order in orders:
            order_tags = order.get('tagIds') or []  # Handle None values
            all_tag_ids.update(str(tag) for tag in order_tags)
        if all_tag_ids:
            print(f"   {', '.join(sorted(all_tag_ids))}")
        else:
            print("   No tags found in any orders")
        return
    
    print(f"[SUCCESS] Total orders found: {len(filtered_orders)}")
    
    # Debug: Check first order structure
    if filtered_orders:
        first_order = filtered_orders[0]
        print("\n[DEBUG] Debug - First order structure:")
        print(f"   Order Number: {first_order.get('orderNumber', 'N/A')}")
        print(f"   Has 'items' field: {'items' in first_order}")
        print(f"   Has 'lineItems' field: {'lineItems' in first_order}")
        if 'items' in first_order:
            items = first_order.get('items', [])
            print(f"   Items count: {len(items)}")
            if items:
                print(f"   First item: {items[0]}")
        if 'lineItems' in first_order:
            line_items = first_order.get('lineItems', [])
            print(f"   LineItems count: {len(line_items)}")
            if line_items:
                print(f"   First lineItem: {line_items[0]}")
        print()
    
    # Display summary
    print("\n[SUMMARY] Order Summary:")
    print("-" * 50)
    for i, order in enumerate(filtered_orders[:10], 1):  # Show first 10 orders
        order_number = order.get('orderNumber', 'N/A')
        customer_name = order.get('customerName', 'N/A')
        amount_paid = order.get('amountPaid', 0)
        order_date = order.get('orderDate', 'N/A')
        
        print(f"{i:2d}. Order #{order_number}")
        print(f"    Customer: {customer_name}")
        print(f"    Amount: ${amount_paid}")
        print(f"    Date: {order_date}")
        print()
    
    if len(filtered_orders) > 10:
        print(f"    ... and {len(filtered_orders) - 10} more orders")
        print()
    
    # Export to all formats
    print("[EXPORT] Exporting orders...")
    try:
        # Create output folder
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_folder = str(tag_output_dir(f"Tag_{tag_id}_Orders_{timestamp}"))
        print(f"[FOLDER] Created output folder: {output_folder}")
        
        # Create files in the folder
        json_filename = os.path.join(output_folder, f"tag_{tag_id}_awaiting_orders_{timestamp}.json")
        detailed_csv_filename = os.path.join(output_folder, f"tag_{tag_id}_awaiting_detailed_{timestamp}.csv")
        packing_list_filename = os.path.join(output_folder, f"packing_list_tag_{tag_id}_awaiting_{timestamp}.csv")
        
        # Export JSON
        print("[JSON] Creating JSON file...")
        with open(json_filename, 'w', encoding='utf-8') as jsonfile:
            json.dump(filtered_orders, jsonfile, indent=2, ensure_ascii=False, default=str)
        
        # Export detailed CSV
        print("[CSV] Creating detailed CSV...")
        detailed_csv_file = shipstation.export_orders_to_csv(filtered_orders, detailed_csv_filename)
        
        print("[STOCK] Loading stock levels (see config.py FTP_LOCAL_FILE)...")
        stock_levels = load_stock_levels()

        print("[PACKS] Loading Packs Database for component mapping...")
        packs_map = load_packs_database()
        pack_names_map = load_pack_names()
        print(f"[PACKS] Packs map entries: {len(packs_map)}; Pack Names: {len(pack_names_map)}")

        custom_label_map, labels_missing_stock_id = load_custom_label_stock_map(log=print)

        process_no = get_process_no_for_tag(tag_id)
        if process_no:
            print(f"[INFO] Process No for Tag {tag_id}: {process_no}")
        else:
            print(f"[WARNING] Process No not found for Tag {tag_id}. Will fallback in EDI order-id.")

        print("[PACKING] Creating packing list with stock validation (pack-aware)...")
        in_stock_items, out_of_stock_items, not_found_items = validate_orders_stock(
            filtered_orders,
            tag_id,
            process_no,
            stock_levels,
            packs_map,
            pack_names_map,
            custom_label_map,
            labels_missing_stock_id=labels_missing_stock_id,
        )

        write_packing_list_csv(packing_list_filename, in_stock_items)

        issues_filename = None
        if out_of_stock_items or not_found_items:
            issues_filename = os.path.join(
                output_folder, f"stock_issues_tag_{tag_id}_{timestamp}.csv"
            )
            print(f"[ISSUES] Creating stock issues list: {issues_filename}")
            write_stock_issues_csv(issues_filename, out_of_stock_items, not_found_items)
            print(
                f"[WARNING] {len(not_found_items)} not found, "
                f"{len(out_of_stock_items)} out of stock -> {issues_filename}"
            )

        print(f"[SUCCESS] Packing list created with {len(in_stock_items)} in-stock items")

        print("[EDI] Creating EDI orders file...")
        edi_orders_filename = os.path.join(output_folder, f"edi_orders_tag_{tag_id}_{timestamp}.csv")
        write_edi_orders_csv(edi_orders_filename, in_stock_items, process_no)
        print(f"[SUCCESS] EDI orders file created: {edi_orders_filename}")

        print("\n[SUCCESS] Export completed!")
        print(f"[FOLDER] All files saved in folder: {os.path.abspath(output_folder)}")
        print(f"[JSON] JSON: {os.path.basename(json_filename)}")
        print(f"[CSV] Detailed CSV: {os.path.basename(detailed_csv_file)}")
        print(f"[PACKING] Packing List (In Stock): {os.path.basename(packing_list_filename)}")
        print(f"[EDI] EDI Orders File: {os.path.basename(edi_orders_filename)}")
        if issues_filename:
            print(f"[ISSUES] Stock Issues: {os.path.basename(issues_filename)}")
        print(f"[TOTAL] Total orders: {len(filtered_orders)}")

        # Generate PDF packing slips for EDI orders only (fully in-stock)
        print("\n[PDF] Generating PDF packing slips (EDI / in-stock orders only)...")
        if not in_stock_items:
            print("[PDF] Skipped — no in-stock (EDI) orders to generate packing slips for.")
        else:
            try:
                pdf_source_filename = os.path.join(
                    output_folder, f"pdf_packing_tag_{tag_id}_awaiting_{timestamp}.csv"
                )
                write_packing_list_csv(
                    pdf_source_filename,
                    rows_for_pdf_slips(
                        in_stock_items,
                        [],
                        [],
                        packs_map=packs_map,
                        pack_names_map=pack_names_map,
                    ),
                )
                pdf_output_path = os.path.join(output_folder, pdf_filename_for_tag(tag_id, process_no))
                if generate_packing_slips_for_tag(pdf_source_filename, tag_id, pdf_output_path):
                    print(
                        f"[SUCCESS] PDF generation completed! Saved to: {os.path.abspath(pdf_output_path)}"
                    )
                else:
                    print(
                        "[WARNING] PDF was not created (no packing-slip rows). "
                        "Check that EDI orders have line items."
                    )
            except Exception as e:
                print(f"[WARNING] PDF generation failed: {e}")
                print("   The CSV files were created successfully, but PDF generation encountered an issue.")

        print()
        print(
            format_run_summary(
                tag_label=str(tag_id),
                orders_processed=len(filtered_orders),
                in_stock_items=in_stock_items,
                out_of_stock_items=out_of_stock_items,
                not_found_items=not_found_items,
                issues_filename=issues_filename,
            )
        )
    except Exception as e:
        print(f"[ERROR] Error exporting orders: {e}")
        print("   Please check if you have write permissions in the current directory")
        sys.exit(1)
    
    print("\n[SUCCESS] Done!")

if __name__ == "__main__":
    main()
